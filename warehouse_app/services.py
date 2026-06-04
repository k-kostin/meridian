import calendar
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import DecimalField, Q, Sum
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import DocumentStatus, InventoryDocument, Item, StockDocumentLine
from .models import StockDocumentType, Warehouse


PRESENTATION_CONSOLIDATED = "consolidated"
PRESENTATION_BY_WAREHOUSE = "by_warehouse"
PRESENTATION_LABELS = {
    PRESENTATION_CONSOLIDATED: "Сводно по складам",
    PRESENTATION_BY_WAREHOUSE: "С разбивкой по складам",
}
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="30404D")
INVALID_SHEET_TITLE_CHARS = '[]:*?/\\'
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
QUANTITY_OUTPUT_FIELD = DecimalField(max_digits=14, decimal_places=3)


@dataclass
class WorkbookExport:
    buffer: BytesIO
    filename: str
    content_type: str = XLSX_CONTENT_TYPE


@dataclass(frozen=True)
class MovementEvent:
    operation_date: date
    document_id: int
    document_number: str
    document_type: str
    document_type_label: str
    warehouse_id: int
    warehouse_name: str
    warehouse_code: str
    item_id: int
    item_sku: str
    item_name: str
    unit_code: str
    unit_precision: int
    quantity: Decimal
    comment: str
    line_id: int


def normalize_presentation(value, *, default=PRESENTATION_CONSOLIDATED):
    if value == PRESENTATION_BY_WAREHOUSE:
        return PRESENTATION_BY_WAREHOUSE
    if value == PRESENTATION_CONSOLIDATED:
        return PRESENTATION_CONSOLIDATED
    return default


def presentation_label(value):
    return PRESENTATION_LABELS[normalize_presentation(value)]


def _balance_value_fields(presentation):
    fields = [
        "item_id",
        "item__sku",
        "item__name",
        "item__unit__code",
        "item__unit__display_precision",
    ]
    if presentation == PRESENTATION_BY_WAREHOUSE:
        return [
            "document__warehouse_id",
            "document__warehouse__name",
            "document__warehouse__code",
            *fields,
        ]
    return fields


def _balance_ordering(presentation):
    if presentation == PRESENTATION_BY_WAREHOUSE:
        return ("document__warehouse__name", "item__name", "item__sku")
    return ("item__name", "item__sku")


def _row_key_from_balance(row, presentation):
    if presentation == PRESENTATION_BY_WAREHOUSE:
        return (row["document__warehouse_id"], row["item_id"])
    return row["item_id"]


def _row_key_from_line(line, presentation):
    if presentation == PRESENTATION_BY_WAREHOUSE:
        return (line.document.warehouse_id, line.item_id)
    return line.item_id


def _row_key_from_event(event: MovementEvent, presentation):
    if presentation == PRESENTATION_BY_WAREHOUSE:
        return (event.warehouse_id, event.item_id)
    return event.item_id


def _base_row(
    *,
    item_id,
    sku,
    name,
    unit_code,
    unit_precision,
    warehouse_id=None,
    warehouse_name="",
    warehouse_code="",
):
    return {
        "item_id": item_id,
        "sku": sku,
        "name": name,
        "unit": unit_code,
        "unit_precision": unit_precision,
        "warehouse_id": warehouse_id,
        "warehouse_name": warehouse_name,
        "warehouse_code": warehouse_code,
        "opening": Decimal("0"),
        "incoming": Decimal("0"),
        "outgoing": Decimal("0"),
        "net": Decimal("0"),
        "closing": Decimal("0"),
        "has_movement": False,
    }


def _base_row_from_balance(row, presentation):
    return _base_row(
        item_id=row["item_id"],
        sku=row["item__sku"],
        name=row["item__name"],
        unit_code=row["item__unit__code"],
        unit_precision=row["item__unit__display_precision"],
        warehouse_id=row.get("document__warehouse_id"),
        warehouse_name=row.get("document__warehouse__name", "") if presentation == PRESENTATION_BY_WAREHOUSE else "",
        warehouse_code=row.get("document__warehouse__code", "") if presentation == PRESENTATION_BY_WAREHOUSE else "",
    )


def _base_row_from_line(line, presentation):
    return _base_row(
        item_id=line.item_id,
        sku=line.item.sku,
        name=line.item.name,
        unit_code=line.item.unit.code,
        unit_precision=line.item.unit.display_precision,
        warehouse_id=line.document.warehouse_id,
        warehouse_name=line.document.warehouse.name if presentation == PRESENTATION_BY_WAREHOUSE else "",
        warehouse_code=line.document.warehouse.code if presentation == PRESENTATION_BY_WAREHOUSE else "",
    )


def _base_row_from_event(event: MovementEvent, presentation):
    return _base_row(
        item_id=event.item_id,
        sku=event.item_sku,
        name=event.item_name,
        unit_code=event.unit_code,
        unit_precision=event.unit_precision,
        warehouse_id=event.warehouse_id,
        warehouse_name=event.warehouse_name if presentation == PRESENTATION_BY_WAREHOUSE else "",
        warehouse_code=event.warehouse_code if presentation == PRESENTATION_BY_WAREHOUSE else "",
    )


def _compute_totals_by_unit(rows):
    """Aggregate rows by unit code. Returns list sorted by unit code.

    Never mixes quantities across different units — each entry covers
    exactly one unit so the sums are dimensionally homogeneous.
    """
    by_unit = {}
    for row in rows:
        u = row["unit"]
        if u not in by_unit:
            by_unit[u] = {
                "unit": u,
                "unit_precision": row.get("unit_precision") or 0,
                "opening": Decimal("0"),
                "incoming": Decimal("0"),
                "outgoing": Decimal("0"),
                "net": Decimal("0"),
                "closing": Decimal("0"),
            }
        for key in ("opening", "incoming", "outgoing", "net", "closing"):
            by_unit[u][key] += row[key]
    return sorted(by_unit.values(), key=lambda t: t["unit"])


def _build_grouped_rows(rows):
    """Group rows by warehouse_id (stable PK), computing per-unit subtotals per group.

    Uses warehouse_id as the group key so two warehouses with identical names
    remain separate groups. warehouse_label includes the code in parentheses
    when two or more warehouses share the same name.
    """
    groups = {}
    group_order = []
    for row in rows:
        wid = row["warehouse_id"]
        if wid not in groups:
            groups[wid] = {
                "warehouse_id": wid,
                "warehouse_name": row["warehouse_name"],
                "warehouse_code": row.get("warehouse_code", ""),
                "rows": [],
            }
            group_order.append(wid)
        groups[wid]["rows"].append(row)

    names = [groups[wid]["warehouse_name"] for wid in group_order]
    show_codes = len(names) != len(set(names))

    return [
        {
            "warehouse_id": wid,
            "warehouse_name": groups[wid]["warehouse_name"],
            "warehouse_code": groups[wid]["warehouse_code"],
            "warehouse_label": (
                f'{groups[wid]["warehouse_name"]} ({groups[wid]["warehouse_code"]})'
                if show_codes
                else groups[wid]["warehouse_name"]
            ),
            "rows": groups[wid]["rows"],
            "subtotals_by_unit": _compute_totals_by_unit(groups[wid]["rows"]),
        }
        for wid in group_order
    ]


def _excel_quantity_format(places):
    places = max(int(places or 0), 0)
    if places == 0:
        return "0"
    return "0." + ("0" * places)


def _apply_header(sheet):
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _set_quantity_cell_format(sheet, row_number, column_number, places):
    sheet.cell(row=row_number, column=column_number).number_format = _excel_quantity_format(places)


def _set_date_cell_format(sheet, row_number, column_number):
    sheet.cell(row=row_number, column=column_number).number_format = "DD.MM.YYYY"


def _set_month_cell_format(sheet, row_number, column_number):
    sheet.cell(row=row_number, column=column_number).number_format = "MM.YYYY"


def _warehouse_label(warehouse):
    return warehouse.name if warehouse is not None else "Все склады"


def _safe_sheet_title(workbook, raw_title):
    cleaned = "".join("_" if char in INVALID_SHEET_TITLE_CHARS else char for char in str(raw_title or "Лист"))
    cleaned = cleaned.strip() or "Лист"
    candidate = cleaned[:31]
    suffix_index = 2
    while candidate in workbook.sheetnames:
        suffix = f" {suffix_index}"
        candidate = f"{cleaned[: max(31 - len(suffix), 1)]}{suffix}"
        suffix_index += 1
    return candidate


def _value_label(value):
    if value in (None, ""):
        return "-"
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    return str(value)


def _apply_balance_warehouse_labels(rows, presentation):
    if presentation != PRESENTATION_BY_WAREHOUSE or not rows:
        return rows
    warehouse_names_by_id = {
        row["document__warehouse_id"]: row["document__warehouse__name"]
        for row in rows
    }
    names = list(warehouse_names_by_id.values())
    show_codes = len(names) != len(set(names))
    for row in rows:
        warehouse_name = row["document__warehouse__name"]
        warehouse_code = row.get("document__warehouse__code", "")
        row["warehouse_label"] = f"{warehouse_name} ({warehouse_code})" if show_codes else warehouse_name
    return rows


def _aggregate_balance_rows(queryset, presentation):
    rows = list(
        queryset.values(*_balance_value_fields(presentation))
        .annotate(
            incoming_total=Coalesce(
                Sum("quantity", filter=Q(quantity__gt=0)),
                Decimal("0"),
                output_field=QUANTITY_OUTPUT_FIELD,
            ),
            outgoing_signed=Coalesce(
                Sum("quantity", filter=Q(quantity__lt=0)),
                Decimal("0"),
                output_field=QUANTITY_OUTPUT_FIELD,
            ),
            quantity=Coalesce(
                Sum("quantity"),
                Decimal("0"),
                output_field=QUANTITY_OUTPUT_FIELD,
            ),
        )
        .order_by(*_balance_ordering(presentation))
    )
    for row in rows:
        row["outgoing_total"] = abs(row.pop("outgoing_signed"))
    return rows


def _balance_row_from_event(event: MovementEvent, presentation):
    row = {
        "item_id": event.item_id,
        "item__sku": event.item_sku,
        "item__name": event.item_name,
        "item__unit__code": event.unit_code,
        "item__unit__display_precision": event.unit_precision,
        "incoming_total": Decimal("0"),
        "outgoing_total": Decimal("0"),
        "quantity": Decimal("0"),
    }
    if presentation == PRESENTATION_BY_WAREHOUSE:
        row.update(
            {
                "document__warehouse_id": event.warehouse_id,
                "document__warehouse__name": event.warehouse_name,
                "document__warehouse__code": event.warehouse_code,
            }
        )
    return row


def _zero_balance_row(item, *, warehouse=None, presentation=PRESENTATION_CONSOLIDATED):
    row = {
        "item_id": item.id,
        "item__sku": item.sku,
        "item__name": item.name,
        "item__unit__code": item.unit.code,
        "item__unit__display_precision": item.unit.display_precision,
        "incoming_total": Decimal("0"),
        "outgoing_total": Decimal("0"),
        "quantity": Decimal("0"),
    }
    if presentation == PRESENTATION_BY_WAREHOUSE and warehouse is not None:
        row.update(
            {
                "document__warehouse_id": warehouse.id,
                "document__warehouse__name": warehouse.name,
                "document__warehouse__code": warehouse.code,
            }
        )
    return row


def _expand_zero_balance_rows(rows, warehouse, presentation):
    row_map = {_row_key_from_balance(row, presentation): row for row in rows}
    items = list(Item.objects.select_related("unit").order_by("name", "sku"))
    expanded = []
    if presentation == PRESENTATION_BY_WAREHOUSE:
        warehouses = [warehouse] if warehouse is not None else list(
            Warehouse.objects.order_by("name", "code")
        )
        for warehouse_obj in warehouses:
            for item in items:
                key = (warehouse_obj.id, item.id)
                expanded.append(
                    row_map.get(
                        key,
                        _zero_balance_row(item, warehouse=warehouse_obj, presentation=presentation),
                    )
                )
    else:
        for item in items:
            expanded.append(row_map.get(item.id, _zero_balance_row(item, presentation=presentation)))
    return expanded


def filter_balance_rows(rows, query, presentation):
    query = (query or "").strip().lower()
    if not query:
        return rows

    filtered = []
    for row in rows:
        haystack = [
            row.get("item__sku", ""),
            row.get("item__name", ""),
            row.get("item__unit__code", ""),
        ]
        if presentation == PRESENTATION_BY_WAREHOUSE:
            haystack.extend(
                [
                    row.get("document__warehouse__name", ""),
                    row.get("document__warehouse__code", ""),
                    row.get("warehouse_label", ""),
                ]
            )
        if query in " ".join(str(part).lower() for part in haystack):
            filtered.append(row)
    return filtered


def get_balance_rows(warehouse=None, as_of_date=None, presentation=PRESENTATION_BY_WAREHOUSE, include_zero=False):
    presentation = normalize_presentation(presentation, default=PRESENTATION_BY_WAREHOUSE)
    row_map = {}
    for event in get_movement_rows(warehouse=warehouse, date_to=as_of_date):
        key = _row_key_from_event(event, presentation)
        row = row_map.setdefault(key, _balance_row_from_event(event, presentation))
        if event.quantity > 0:
            row["incoming_total"] += event.quantity
        elif event.quantity < 0:
            row["outgoing_total"] += abs(event.quantity)
        row["quantity"] += event.quantity

    rows = list(row_map.values())
    if include_zero:
        rows = _expand_zero_balance_rows(rows, warehouse, presentation)
    else:
        rows = [row for row in rows if row["quantity"] != 0]
    if presentation == PRESENTATION_BY_WAREHOUSE:
        rows.sort(key=lambda row: (row.get("document__warehouse__name", ""), row["item__name"], row["item__sku"]))
    else:
        rows.sort(key=lambda row: (row["item__name"], row["item__sku"]))
    return _apply_balance_warehouse_labels(rows, presentation)


def get_balance_map(warehouse=None, as_of_date=None, presentation=PRESENTATION_CONSOLIDATED):
    presentation = normalize_presentation(presentation)
    balances = {}
    for row in get_balance_rows(warehouse=warehouse, as_of_date=as_of_date, presentation=presentation):
        key = _row_key_from_balance(row, presentation)
        balances[key] = balances.get(key, Decimal("0")) + row["quantity"]
    return balances


def _movement_bucket_template():
    return {
        "incoming": Decimal("0"),
        "outgoing": Decimal("0"),
        "net": Decimal("0"),
    }


def _ledger_row_meta(row, presentation):
    meta = {
        "item_id": row["item_id"],
        "sku": row["item__sku"],
        "name": row["item__name"],
        "unit": row["item__unit__code"],
        "unit_precision": row["item__unit__display_precision"],
    }
    if presentation == PRESENTATION_BY_WAREHOUSE:
        meta.update(
            {
                "warehouse_id": row["document__warehouse_id"],
                "warehouse_name": row["document__warehouse__name"],
                "warehouse_code": row["document__warehouse__code"],
                "warehouse_label": row.get("warehouse_label", row["document__warehouse__name"]),
            }
        )
    return meta


def _month_start(value: date) -> date:
    return value.replace(day=1)


def _month_end(value: date) -> date:
    return value.replace(day=calendar.monthrange(value.year, value.month)[1])


def _iter_month_starts(period_start: date, period_end: date):
    current = _month_start(period_start)
    end = _month_start(period_end)
    while current <= end:
        yield current
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1, day=1)
        else:
            current = current.replace(month=current.month + 1, day=1)


def build_daily_ledger(
    warehouse=None,
    period_start=None,
    period_end=None,
    presentation=PRESENTATION_CONSOLIDATED,
):
    presentation = normalize_presentation(presentation)
    if period_start is None or period_end is None:
        resolved = resolve_period("month")
        period_start = resolved["start"]
        period_end = resolved["end"]

    opening_rows = get_balance_rows(
        warehouse=warehouse,
        as_of_date=period_start - timedelta(days=1),
        presentation=presentation,
        include_zero=True,
    )
    key_order = [_row_key_from_balance(row, presentation) for row in opening_rows]
    meta_map = {key: _ledger_row_meta(row, presentation) for key, row in zip(key_order, opening_rows)}
    current_balances = {key: row["quantity"] for key, row in zip(key_order, opening_rows)}

    movement_rows = get_movement_rows(warehouse=warehouse, date_from=period_start, date_to=period_end)
    movement_map = {}
    document_numbers = set()
    line_keys = set()

    for movement in movement_rows:
        line_keys.add((movement.document_id, movement.line_id))
        document_numbers.add(movement.document_number)
        key = _row_key_from_event(movement, presentation)
        bucket_key = (movement.operation_date, key)
        bucket = movement_map.setdefault(bucket_key, _movement_bucket_template())
        if movement.quantity > 0:
            bucket["incoming"] += movement.quantity
        elif movement.quantity < 0:
            bucket["outgoing"] += abs(movement.quantity)
        bucket["net"] += movement.quantity

    rows = []
    current_date = period_start
    while current_date <= period_end:
        for key in key_order:
            movement = movement_map.get((current_date, key), _movement_bucket_template())
            opening = current_balances.get(key, Decimal("0"))
            closing = opening + movement["net"]
            rows.append(
                {
                    **meta_map[key],
                    "date": current_date,
                    "opening": opening,
                    "incoming": movement["incoming"],
                    "outgoing": movement["outgoing"],
                    "net": movement["net"],
                    "closing": closing,
                }
            )
            current_balances[key] = closing
        current_date += timedelta(days=1)

    return {
        "rows": rows,
        "presentation": presentation,
        "presentation_label": presentation_label(presentation),
        "period": {
            "start": period_start,
            "end": period_end,
            "label": f"Дни за {period_start.strftime('%m.%Y')}",
        },
        "summary": {
            "rows_count": len(rows),
            "days_count": (period_end - period_start).days + 1,
            "documents_count": len(document_numbers),
            "lines_count": len(line_keys),
        },
    }


def build_monthly_ledger(
    warehouse=None,
    period_start=None,
    period_end=None,
    presentation=PRESENTATION_CONSOLIDATED,
):
    presentation = normalize_presentation(presentation)
    if period_start is None or period_end is None:
        today = timezone.localdate()
        period_start = today.replace(day=1)
        period_end = _month_end(today)

    normalized_start = _month_start(period_start)
    normalized_end = _month_end(period_end)
    opening_rows = get_balance_rows(
        warehouse=warehouse,
        as_of_date=normalized_start - timedelta(days=1),
        presentation=presentation,
        include_zero=True,
    )
    key_order = [_row_key_from_balance(row, presentation) for row in opening_rows]
    meta_map = {key: _ledger_row_meta(row, presentation) for key, row in zip(key_order, opening_rows)}
    current_balances = {key: row["quantity"] for key, row in zip(key_order, opening_rows)}

    movement_rows = get_movement_rows(
        warehouse=warehouse,
        date_from=normalized_start,
        date_to=normalized_end,
    )
    movement_map = {}
    document_numbers = set()
    line_keys = set()

    for movement in movement_rows:
        line_keys.add((movement.document_id, movement.line_id))
        document_numbers.add(movement.document_number)
        key = _row_key_from_event(movement, presentation)
        bucket_key = (_month_start(movement.operation_date), key)
        bucket = movement_map.setdefault(bucket_key, _movement_bucket_template())
        if movement.quantity > 0:
            bucket["incoming"] += movement.quantity
        elif movement.quantity < 0:
            bucket["outgoing"] += abs(movement.quantity)
        bucket["net"] += movement.quantity

    rows = []
    months = list(_iter_month_starts(normalized_start, normalized_end))
    for month_start in months:
        for key in key_order:
            movement = movement_map.get((month_start, key), _movement_bucket_template())
            opening = current_balances.get(key, Decimal("0"))
            closing = opening + movement["net"]
            rows.append(
                {
                    **meta_map[key],
                    "month_start": month_start,
                    "month_end": _month_end(month_start),
                    "month_label": month_start.strftime("%m.%Y"),
                    "opening": opening,
                    "incoming": movement["incoming"],
                    "outgoing": movement["outgoing"],
                    "net": movement["net"],
                    "closing": closing,
                }
            )
            current_balances[key] = closing

    return {
        "rows": rows,
        "presentation": presentation,
        "presentation_label": presentation_label(presentation),
        "period": {
            "start": normalized_start,
            "end": normalized_end,
            "label": f"Месяцы {normalized_start.strftime('%m.%Y')} - {normalized_end.strftime('%m.%Y')}",
        },
        "summary": {
            "rows_count": len(rows),
            "months_count": len(months),
            "documents_count": len(document_numbers),
            "lines_count": len(line_keys),
        },
    }


def _posted_movement_lines(warehouse=None, date_from=None, date_to=None, document_type=None, status=None):
    target_status = status or DocumentStatus.POSTED
    queryset = (
        StockDocumentLine.objects.filter(document__status=target_status)
        .select_related("document", "item", "item__unit", "document__warehouse", "document__destination_warehouse")
        .order_by("-document__operation_date", "-document__id", "id")
    )
    if document_type:
        queryset = queryset.filter(document__document_type=document_type)
    if date_from is not None:
        queryset = queryset.filter(document__operation_date__gte=date_from)
    if date_to is not None:
        queryset = queryset.filter(document__operation_date__lte=date_to)
    return queryset


def _movement_event(line: StockDocumentLine, warehouse, quantity: Decimal):
    return MovementEvent(
        operation_date=line.document.operation_date,
        document_id=line.document_id,
        document_number=line.document.number,
        document_type=line.document.document_type,
        document_type_label=line.document.get_document_type_display(),
        warehouse_id=warehouse.id,
        warehouse_name=warehouse.name,
        warehouse_code=warehouse.code,
        item_id=line.item_id,
        item_sku=line.item.sku,
        item_name=line.item.name,
        unit_code=line.item.unit.code,
        unit_precision=line.item.unit.display_precision,
        quantity=quantity,
        comment=line.comment or line.document.comment,
        line_id=line.id,
    )


def get_movement_rows(warehouse=None, date_from=None, date_to=None, document_type=None, status=None):
    events = []
    for line in _posted_movement_lines(date_from=date_from, date_to=date_to, document_type=document_type, status=status):
        if line.document.document_type == StockDocumentType.TRANSFER:
            source_quantity = -abs(line.quantity)
            destination_quantity = abs(line.quantity)
            source_event = _movement_event(line, line.document.warehouse, source_quantity)
            destination_event = _movement_event(line, line.document.destination_warehouse, destination_quantity)
            for event in (source_event, destination_event):
                if warehouse is None or event.warehouse_id == warehouse.id:
                    events.append(event)
            continue

        event = _movement_event(line, line.document.warehouse, line.quantity)
        if warehouse is None or event.warehouse_id == warehouse.id:
            events.append(event)
    return events


def resolve_period(mode, anchor_date=None, date_from=None, date_to=None):
    today = timezone.localdate()
    mode = mode or "day"
    anchor_date = anchor_date or today

    if mode == "month":
        start = anchor_date.replace(day=1)
        last_day = calendar.monthrange(anchor_date.year, anchor_date.month)[1]
        end = anchor_date.replace(day=last_day)
        label = f"Месяц {start.strftime('%m.%Y')}"
    elif mode == "year":
        start = anchor_date.replace(month=1, day=1)
        end = anchor_date.replace(month=12, day=31)
        label = f"Год {anchor_date.year}"
    elif mode == "custom":
        start = date_from or today
        end = date_to or start
        if start > end:
            start, end = end, start
        label = f"Период {start.strftime('%d.%m.%Y')} - {end.strftime('%d.%m.%Y')}"
    else:
        start = anchor_date
        end = anchor_date
        label = f"Дата {anchor_date.strftime('%d.%m.%Y')}"

    return {
        "mode": mode,
        "start": start,
        "end": end,
        "label": label,
        "is_single_day": start == end,
    }


def build_period_report(
    warehouse=None,
    period_start=None,
    period_end=None,
    presentation=PRESENTATION_CONSOLIDATED,
):
    presentation = normalize_presentation(presentation)
    if period_start is None or period_end is None:
        resolved = resolve_period("day")
        period_start = resolved["start"]
        period_end = resolved["end"]

    opening_date = period_start - timedelta(days=1)
    opening_rows = get_balance_rows(warehouse=warehouse, as_of_date=opening_date, presentation=presentation)
    closing_rows = get_balance_rows(warehouse=warehouse, as_of_date=period_end, presentation=presentation)
    opening_map = {_row_key_from_balance(row, presentation): row["quantity"] for row in opening_rows}
    closing_map = {_row_key_from_balance(row, presentation): row["quantity"] for row in closing_rows}

    movement_qs = get_movement_rows(warehouse=warehouse, date_from=period_start, date_to=period_end)
    row_map = {
        _row_key_from_balance(row, presentation): _base_row_from_balance(row, presentation) for row in opening_rows
    }
    for row in closing_rows:
        row_map.setdefault(_row_key_from_balance(row, presentation), _base_row_from_balance(row, presentation))
    document_numbers = set()
    line_keys = set()

    for movement in movement_qs:
        line_keys.add((movement.document_id, movement.line_id))
        document_numbers.add(movement.document_number)
        key = _row_key_from_event(movement, presentation)
        row = row_map.setdefault(key, _base_row_from_event(movement, presentation))
        if movement.quantity > 0:
            row["incoming"] += movement.quantity
        elif movement.quantity < 0:
            row["outgoing"] += abs(movement.quantity)
        row["net"] += movement.quantity
        row["has_movement"] = row["incoming"] != 0 or row["outgoing"] != 0

    rows = []
    all_keys = set(opening_map) | set(closing_map) | set(row_map)
    for key in all_keys:
        row = row_map[key].copy()
        row["opening"] = opening_map.get(key, Decimal("0"))
        row["closing"] = closing_map.get(key, Decimal("0"))
        row["has_movement"] = row["incoming"] != 0 or row["outgoing"] != 0
        rows.append(row)

    if presentation == PRESENTATION_BY_WAREHOUSE:
        rows.sort(key=lambda row: (row["warehouse_name"], row["name"], row["sku"]))
    else:
        rows.sort(key=lambda row: (row["name"], row["sku"]))

    grand_total_by_unit = _compute_totals_by_unit(rows)
    grouped_rows = _build_grouped_rows(rows) if presentation == PRESENTATION_BY_WAREHOUSE else None

    # Propagate warehouse_label from groups back onto individual rows so the
    # label is available in templates and exports without extra lookups.
    if grouped_rows is not None:
        for group in grouped_rows:
            for row in group["rows"]:
                row["warehouse_label"] = group["warehouse_label"]

    return {
        "rows": rows,
        "grand_total_by_unit": grand_total_by_unit,
        "grouped_rows": grouped_rows,
        "presentation": presentation,
        "presentation_label": presentation_label(presentation),
        "summary": {
            "positions_count": len(rows),
            "rows_count": len(rows),
            "movement_positions_count": sum(1 for row in rows if row["has_movement"]),
            "documents_count": len(document_numbers),
            "lines_count": len(line_keys),
            "opening_positions_count": sum(1 for row in rows if row["opening"] != 0),
            "closing_positions_count": sum(1 for row in rows if row["closing"] != 0),
        },
    }


def _format_sheet(worksheet, title, freeze_panes="A2"):
    worksheet.freeze_panes = freeze_panes
    worksheet.sheet_view.showGridLines = False
    worksheet.title = title


def _append_metadata_sheet(workbook, metadata_rows):
    sheet = workbook.create_sheet("Параметры", 0)
    _format_sheet(sheet, "Параметры", freeze_panes="A2")
    sheet.append(["Параметр", "Значение"])
    _apply_header(sheet)
    for label, value in metadata_rows:
        sheet.append([label, _value_label(value)])
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 42


def _workbook_export(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return WorkbookExport(buffer=buffer, filename=filename)


def export_items_xlsx():
    workbook = Workbook()
    sheet = workbook.active
    _format_sheet(sheet, "Номенклатура")
    items = list(Item.objects.select_related("unit").order_by("name"))
    _append_metadata_sheet(
        workbook,
        [
            ("Отчет", "Номенклатура"),
            ("Сформировано", timezone.localtime()),
            ("Строк номенклатуры", len(items)),
        ],
    )
    headers = ["Артикул", "Наименование", "Единица", "Активна", "Комментарий"]
    sheet.append(headers)
    _apply_header(sheet)
    for item in items:
        sheet.append([item.sku, item.name, item.unit.code, "Да" if item.is_active else "Нет", item.notes])
    workbook.active = 1
    return _workbook_export(workbook, "items.xlsx")


def export_balances_xlsx(warehouse=None, presentation=PRESENTATION_BY_WAREHOUSE, query="", include_zero=False):
    presentation = normalize_presentation(presentation, default=PRESENTATION_BY_WAREHOUSE)
    rows = filter_balance_rows(
        get_balance_rows(warehouse=warehouse, presentation=presentation, include_zero=include_zero),
        query=query,
        presentation=presentation,
    )
    workbook = Workbook()
    sheet = workbook.active
    _format_sheet(sheet, "Остатки")
    _append_metadata_sheet(
        workbook,
        [
            ("Отчет", "Текущие остатки"),
            ("Сформировано", timezone.localtime()),
            ("Фильтр по складу", _warehouse_label(warehouse)),
            ("Представление", presentation_label(presentation)),
            ("Поиск", query or "-"),
            ("Показывать нулевые позиции", "Да" if include_zero else "Нет"),
            ("Строк остатков", len(rows)),
        ],
    )
    headers = ["Артикул", "Наименование", "Единица", "Приход всего", "Уход всего", "Остаток"]
    if presentation == PRESENTATION_BY_WAREHOUSE:
        headers.insert(0, "Склад")
    sheet.append(headers)
    _apply_header(sheet)
    quantity_columns = (5, 6, 7) if presentation == PRESENTATION_BY_WAREHOUSE else (4, 5, 6)
    for row in rows:
        payload = [
            row["item__sku"],
            row["item__name"],
            row["item__unit__code"],
            row["incoming_total"],
            row["outgoing_total"],
            row["quantity"],
        ]
        if presentation == PRESENTATION_BY_WAREHOUSE:
            payload.insert(0, row["warehouse_label"])
        sheet.append(payload)
        for column_number in quantity_columns:
            _set_quantity_cell_format(sheet, sheet.max_row, column_number, row["item__unit__display_precision"])
    workbook.active = 1
    return _workbook_export(workbook, "balances.xlsx")


def export_movements_xlsx(warehouse=None, date_from=None, date_to=None, document_type=None, status=None):
    lines = list(
        get_movement_rows(
            warehouse=warehouse,
            date_from=date_from,
            date_to=date_to,
            document_type=document_type,
            status=status,
        )
    )
    workbook = Workbook()
    sheet = workbook.active
    _format_sheet(sheet, "Движения")
    _append_metadata_sheet(
        workbook,
        [
            ("Отчет", "Движения"),
            ("Сформировано", timezone.localtime()),
            ("Фильтр по складу", _warehouse_label(warehouse)),
            ("Тип документа", dict(StockDocumentType.choices).get(document_type, "Все типы") if document_type else "Все типы"),
            ("Статус документа", dict(DocumentStatus.choices).get(status, "Все статусы") if status else "Все статусы"),
            ("Дата от", date_from),
            ("Дата до", date_to),
            ("Строк движений", len(lines)),
        ],
    )
    headers = ["Дата", "Номер", "Тип", "Склад", "Артикул", "Наименование", "Количество", "Комментарий"]
    sheet.append(headers)
    _apply_header(sheet)
    for line in lines:
        sheet.append(
            [
                line.operation_date,
                line.document_number,
                line.document_type_label,
                line.warehouse_name,
                line.item_sku,
                line.item_name,
                line.quantity,
                line.comment,
            ]
        )
        _set_date_cell_format(sheet, sheet.max_row, 1)
        _set_quantity_cell_format(sheet, sheet.max_row, 7, line.unit_precision)
    workbook.active = 1
    return _workbook_export(workbook, "movements.xlsx")


def export_inventories_xlsx():
    inventory_rows = []
    inventories = InventoryDocument.objects.select_related("warehouse").prefetch_related("lines__item__unit")
    for inventory in inventories.order_by("-inventory_date", "-id"):
        for line in inventory.lines.select_related("item__unit"):
            variance = line.actual_quantity - line.expected_quantity
            inventory_rows.append((inventory, line, variance))

    workbook = Workbook()
    sheet = workbook.active
    _format_sheet(sheet, "Инвентаризации")
    _append_metadata_sheet(
        workbook,
        [
            ("Отчет", "Инвентаризации"),
            ("Сформировано", timezone.localtime()),
            ("Строк выгрузки", len(inventory_rows)),
        ],
    )
    headers = [
        "Дата",
        "Номер инвентаризации",
        "Склад",
        "Номенклатура",
        "Учетное количество",
        "Фактическое количество",
        "Отклонение",
    ]
    sheet.append(headers)
    _apply_header(sheet)
    for inventory, line, variance in inventory_rows:
        sheet.append(
            [
                inventory.inventory_date,
                inventory.number,
                inventory.warehouse.name,
                line.item.name,
                line.expected_quantity,
                line.actual_quantity,
                variance,
            ]
        )
        _set_date_cell_format(sheet, sheet.max_row, 1)
        _set_quantity_cell_format(sheet, sheet.max_row, 5, line.item.unit.display_precision)
        _set_quantity_cell_format(sheet, sheet.max_row, 6, line.item.unit.display_precision)
        _set_quantity_cell_format(sheet, sheet.max_row, 7, line.item.unit.display_precision)
    workbook.active = 1
    return _workbook_export(workbook, "inventories.xlsx")


def export_daily_ledger_xlsx(
    warehouse=None,
    period_start=None,
    period_end=None,
    presentation=PRESENTATION_CONSOLIDATED,
):
    presentation = normalize_presentation(presentation)
    report = build_daily_ledger(
        warehouse=warehouse,
        period_start=period_start,
        period_end=period_end,
        presentation=presentation,
    )
    workbook = Workbook()
    sheet = workbook.active
    _format_sheet(sheet, "Дни")
    _append_metadata_sheet(
        workbook,
        [
            ("Отчет", "Подневный учет"),
            ("Сформировано", timezone.localtime()),
            ("Период", report["period"]["label"]),
            ("Начало периода", report["period"]["start"]),
            ("Конец периода", report["period"]["end"]),
            ("Фильтр по складу", _warehouse_label(warehouse)),
            ("Представление", presentation_label(presentation)),
            ("Строк в отчете", report["summary"]["rows_count"]),
            ("Дней в периоде", report["summary"]["days_count"]),
            ("Строк движений", report["summary"]["lines_count"]),
            ("Документов в периоде", report["summary"]["documents_count"]),
        ],
    )
    headers = ["Дата", "Артикул", "Наименование", "Единица", "На начало дня", "Приход", "Расход", "Дельта", "На конец дня"]
    if presentation == PRESENTATION_BY_WAREHOUSE:
        headers.insert(1, "Склад")
    sheet.append(headers)
    _apply_header(sheet)
    quantity_columns = (6, 7, 8, 9, 10) if presentation == PRESENTATION_BY_WAREHOUSE else (5, 6, 7, 8, 9)
    for row in report["rows"]:
        payload = [
            row["date"],
            row["sku"],
            row["name"],
            row["unit"],
            row["opening"],
            row["incoming"],
            row["outgoing"],
            row["net"],
            row["closing"],
        ]
        if presentation == PRESENTATION_BY_WAREHOUSE:
            payload.insert(1, row["warehouse_label"])
        sheet.append(payload)
        _set_date_cell_format(sheet, sheet.max_row, 1)
        for column_number in quantity_columns:
            _set_quantity_cell_format(sheet, sheet.max_row, column_number, row["unit_precision"])
    workbook.active = 1
    filename = f"days_{report['period']['start'].strftime('%Y-%m')}.xlsx"
    return _workbook_export(workbook, filename)


def export_monthly_ledger_xlsx(
    warehouse=None,
    period_start=None,
    period_end=None,
    presentation=PRESENTATION_CONSOLIDATED,
):
    presentation = normalize_presentation(presentation)
    report = build_monthly_ledger(
        warehouse=warehouse,
        period_start=period_start,
        period_end=period_end,
        presentation=presentation,
    )
    workbook = Workbook()
    sheet = workbook.active
    _format_sheet(sheet, "Месяцы")
    _append_metadata_sheet(
        workbook,
        [
            ("Отчет", "Помесячная динамика"),
            ("Сформировано", timezone.localtime()),
            ("Период", report["period"]["label"]),
            ("Начало периода", report["period"]["start"]),
            ("Конец периода", report["period"]["end"]),
            ("Фильтр по складу", _warehouse_label(warehouse)),
            ("Представление", presentation_label(presentation)),
            ("Строк в отчете", report["summary"]["rows_count"]),
            ("Месяцев в периоде", report["summary"]["months_count"]),
            ("Строк движений", report["summary"]["lines_count"]),
            ("Документов в периоде", report["summary"]["documents_count"]),
        ],
    )
    headers = ["Месяц", "Артикул", "Наименование", "Единица", "На начало месяца", "Приход", "Расход", "Дельта", "На конец месяца"]
    if presentation == PRESENTATION_BY_WAREHOUSE:
        headers.insert(1, "Склад")
    sheet.append(headers)
    _apply_header(sheet)
    quantity_columns = (6, 7, 8, 9, 10) if presentation == PRESENTATION_BY_WAREHOUSE else (5, 6, 7, 8, 9)
    for row in report["rows"]:
        payload = [
            row["month_start"],
            row["sku"],
            row["name"],
            row["unit"],
            row["opening"],
            row["incoming"],
            row["outgoing"],
            row["net"],
            row["closing"],
        ]
        if presentation == PRESENTATION_BY_WAREHOUSE:
            payload.insert(1, row["warehouse_label"])
        sheet.append(payload)
        _set_month_cell_format(sheet, sheet.max_row, 1)
        for column_number in quantity_columns:
            _set_quantity_cell_format(sheet, sheet.max_row, column_number, row["unit_precision"])
    workbook.active = 1
    filename = (
        f"months_{report['period']['start'].strftime('%Y-%m')}_{report['period']['end'].strftime('%Y-%m')}.xlsx"
    )
    return _workbook_export(workbook, filename)


def export_period_analysis_xlsx(
    warehouse=None,
    period_start=None,
    period_end=None,
    label="period",
    presentation=PRESENTATION_CONSOLIDATED,
    mode_label="День",
):
    presentation = normalize_presentation(presentation)
    report = build_period_report(
        warehouse=warehouse,
        period_start=period_start,
        period_end=period_end,
        presentation=presentation,
    )
    workbook = Workbook()
    sheet = workbook.active
    _format_sheet(sheet, "Аналитика")
    _append_metadata_sheet(
        workbook,
        [
            ("Отчет", "Аналитика периода"),
            ("Сформировано", timezone.localtime()),
            ("Режим периода", mode_label),
            ("Период", label),
            ("Начало периода", period_start),
            ("Конец периода", period_end),
            ("Фильтр по складу", _warehouse_label(warehouse)),
            ("Представление", presentation_label(presentation)),
            ("Строк в отчете", report["summary"]["rows_count"]),
            ("Строк движений", report["summary"]["lines_count"]),
            ("Документов в периоде", report["summary"]["documents_count"]),
            ("Лист «Итоги»", "Итоги по единицам измерения; каждая строка покрывает одну единицу, смешивание исключено."),
            (
                "Доп. листы по складам",
                "Да" if presentation == PRESENTATION_BY_WAREHOUSE and report["grouped_rows"] else "Нет",
            ),
        ],
    )
    headers = ["Артикул", "Наименование", "Единица", "Остаток на начало", "Приход за период", "Расход за период", "Дельта", "Остаток на конец"]
    if presentation == PRESENTATION_BY_WAREHOUSE:
        headers.insert(0, "Склад")
    sheet.append(headers)
    _apply_header(sheet)
    quantity_columns = (5, 6, 7, 8, 9) if presentation == PRESENTATION_BY_WAREHOUSE else (4, 5, 6, 7, 8)
    for row in report["rows"]:
        payload = [
            row["sku"],
            row["name"],
            row["unit"],
            row["opening"],
            row["incoming"],
            row["outgoing"],
            row["net"],
            row["closing"],
        ]
        if presentation == PRESENTATION_BY_WAREHOUSE:
            payload.insert(0, row["warehouse_label"])
        sheet.append(payload)
        for column_number in quantity_columns:
            _set_quantity_cell_format(sheet, sheet.max_row, column_number, row["unit_precision"])

    # "Итоги" sheet: per-unit subtotals (never mixes different units)
    totals_sheet = workbook.create_sheet("Итоги")
    _format_sheet(totals_sheet, "Итоги", freeze_panes="A2")

    if report["grouped_rows"]:
        # by_warehouse: columns — Склад, Ед., На начало, Приход, Расход, Дельта, На конец
        totals_sheet.append(["Склад", "Ед.", "На начало", "Приход", "Расход", "Дельта", "На конец"])
        _apply_header(totals_sheet)
        totals_sheet.column_dimensions["A"].width = 30
        totals_sheet.column_dimensions["B"].width = 8
        for col_letter in ("C", "D", "E", "F", "G"):
            totals_sheet.column_dimensions[col_letter].width = 16

        for group in report["grouped_rows"]:
            for st in group["subtotals_by_unit"]:
                totals_sheet.append([
                    group["warehouse_label"], st["unit"],
                    st["opening"], st["incoming"], st["outgoing"], st["net"], st["closing"],
                ])
                for col_num in range(3, 8):
                    _set_quantity_cell_format(totals_sheet, totals_sheet.max_row, col_num, st["unit_precision"])

        for gt in report["grand_total_by_unit"]:
            totals_sheet.append([
                "Итого по всем складам", gt["unit"],
                gt["opening"], gt["incoming"], gt["outgoing"], gt["net"], gt["closing"],
            ])
            for col_num in range(3, 8):
                _set_quantity_cell_format(totals_sheet, totals_sheet.max_row, col_num, gt["unit_precision"])
            for cell in totals_sheet[totals_sheet.max_row]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
    else:
        # consolidated: columns — Ед., На начало, Приход, Расход, Дельта, На конец
        totals_sheet.append(["Ед.", "На начало", "Приход", "Расход", "Дельта", "На конец"])
        _apply_header(totals_sheet)
        totals_sheet.column_dimensions["A"].width = 8
        for col_letter in ("B", "C", "D", "E", "F"):
            totals_sheet.column_dimensions[col_letter].width = 16

        for gt in report["grand_total_by_unit"]:
            totals_sheet.append([
                gt["unit"],
                gt["opening"], gt["incoming"], gt["outgoing"], gt["net"], gt["closing"],
            ])
            for col_num in range(2, 7):
                _set_quantity_cell_format(totals_sheet, totals_sheet.max_row, col_num, gt["unit_precision"])
            for cell in totals_sheet[totals_sheet.max_row]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL

    if report["grouped_rows"]:
        for group in report["grouped_rows"]:
            warehouse_sheet = workbook.create_sheet(
                _safe_sheet_title(workbook, f"Склад {group['warehouse_label']}")
            )
            _format_sheet(warehouse_sheet, warehouse_sheet.title)
            warehouse_sheet.append(
                ["Артикул", "Наименование", "Единица", "Остаток на начало", "Приход за период", "Расход за период", "Дельта", "Остаток на конец"]
            )
            _apply_header(warehouse_sheet)
            for row in group["rows"]:
                warehouse_sheet.append(
                    [
                        row["sku"],
                        row["name"],
                        row["unit"],
                        row["opening"],
                        row["incoming"],
                        row["outgoing"],
                        row["net"],
                        row["closing"],
                    ]
                )
                for column_number in range(4, 9):
                    _set_quantity_cell_format(warehouse_sheet, warehouse_sheet.max_row, column_number, row["unit_precision"])

            for subtotal in group["subtotals_by_unit"]:
                warehouse_sheet.append(
                    [
                        f"Итого ({group['warehouse_label']})",
                        "",
                        subtotal["unit"],
                        subtotal["opening"],
                        subtotal["incoming"],
                        subtotal["outgoing"],
                        subtotal["net"],
                        subtotal["closing"],
                    ]
                )
                for column_number in range(4, 9):
                    _set_quantity_cell_format(
                        warehouse_sheet,
                        warehouse_sheet.max_row,
                        column_number,
                        subtotal["unit_precision"],
                    )
                for cell in warehouse_sheet[warehouse_sheet.max_row]:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL

    workbook.active = 1
    filename = f"analysis_{period_start.isoformat()}_{period_end.isoformat()}.xlsx"
    return _workbook_export(workbook, filename)
