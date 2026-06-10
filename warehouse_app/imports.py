from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
import re
from typing import BinaryIO

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import InventoryDocument, InventoryLine, InventoryScope, Item, Unit, Warehouse


@dataclass(frozen=True)
class ItemImportRow:
    row_number: int
    sku: str
    name: str
    unit_code: str
    is_active: bool
    comment: str


@dataclass(frozen=True)
class ImportErrorDetail:
    row_number: int
    message: str


@dataclass(frozen=True)
class ItemImportResult:
    rows: list[ItemImportRow]
    errors: list[ImportErrorDetail]


@dataclass(frozen=True)
class ItemImportCommitResult:
    created_count: int
    updated_count: int
    errors: list[ImportErrorDetail]


@dataclass(frozen=True)
class OpeningInventoryImportRow:
    row_number: int
    warehouse_code: str
    sku: str
    actual_quantity: Decimal
    comment: str


@dataclass(frozen=True)
class OpeningInventoryImportResult:
    rows: list[OpeningInventoryImportRow]
    errors: list[ImportErrorDetail]


@dataclass(frozen=True)
class OpeningInventoryImportCommitResult:
    inventory: InventoryDocument | None
    created_lines_count: int
    errors: list[ImportErrorDetail]


REQUIRED_COLUMNS = {
    "Артикул": "Артикул обязателен",
    "Наименование": "Наименование обязательно",
    "Единица": "Единица обязательна",
}
ITEM_COLUMN_ALIASES = {
    "Артикул": ["Артикул", "SKU", "Код", "Код номенклатуры", "Код товара", "Артикул товара"],
    "Наименование": ["Наименование", "Название", "Номенклатура", "Наименование товара", "Наименование позиции", "Товар"],
    "Единица": ["Единица", "Ед.изм.", "Ед изм", "Ед. изм.", "Единица измерения", "ЕИ", "Ед"],
    "Активна": ["Активна", "Активен", "Действует", "Активность"],
    "Комментарий": ["Комментарий", "Примечание"],
}
OPENING_INVENTORY_REQUIRED_COLUMNS = {
    "Склад": "Склад обязателен",
    "Артикул": "Артикул обязателен",
    "Фактическое количество": "Фактическое количество обязательно",
}
OPENING_INVENTORY_COLUMN_ALIASES = {
    "Склад": ["Склад", "Код склада", "Warehouse", "Warehouse code"],
    "Артикул": ["Артикул", "SKU", "Код", "Код номенклатуры", "Код товара", "Артикул товара"],
    "Фактическое количество": ["Фактическое количество", "Количество", "Остаток", "Факт", "Кол-во", "Кол во", "Qty", "Quantity", "Факт. остаток"],
    "Комментарий": ["Комментарий", "Примечание"],
}
ITEM_SHEET_ALIASES = ["Номенклатура", "Товары", "Справочник", "Items"]
OPENING_INVENTORY_SHEET_ALIASES = ["Стартовые остатки", "Остатки", "Остатки склада", "Opening stock"]
ITEM_IMPORT_MODE_CREATE_ONLY = "create_only"
ITEM_IMPORT_MODE_UPDATE_EXISTING = "update_existing"
ITEM_IMPORT_MODES = {ITEM_IMPORT_MODE_CREATE_ONLY, ITEM_IMPORT_MODE_UPDATE_EXISTING}


def _as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized_lookup_key(value) -> str:
    text = _as_text(value).casefold()
    return re.sub(r"[^\w]+", "", text).replace("_", "")


def _as_bool(value) -> bool:
    text = _as_text(value).lower()
    if text in {"нет", "no", "false", "0", "выкл", "off", "disabled"}:
        return False
    return True


def _as_decimal(value) -> tuple[Decimal, bool]:
    text = _as_text(value).replace(",", ".")
    if not text:
        return Decimal("0"), False
    try:
        return Decimal(text), True
    except (InvalidOperation, ValueError):
        return Decimal("0"), False


def _header_map(header_row) -> dict[str, int]:
    return {_normalized_lookup_key(value): index for index, value in enumerate(header_row)}


def _cell(row, headers: dict[str, int], column: str) -> str:
    index = headers.get(_normalized_lookup_key(column))
    if index is None or index >= len(row):
        return ""
    return _as_text(row[index])


def _resolve_column_index(headers: dict[str, int], column: str, aliases: dict[str, list[str]]) -> int | None:
    for candidate in aliases.get(column, [column]):
        index = headers.get(_normalized_lookup_key(candidate))
        if index is not None:
            return index
    return None


def _cell_with_aliases(row, resolved_columns: dict[str, int | None], column: str) -> str:
    index = resolved_columns[column]
    if index is None or index >= len(row):
        return ""
    return _as_text(row[index])


def _resolve_columns(headers: dict[str, int], columns: list[str], aliases: dict[str, list[str]]) -> dict[str, int | None]:
    return {
        column: _resolve_column_index(headers, column, aliases)
        for column in columns
    }


def _required_cell(row, resolved_columns: dict[str, int | None], column: str) -> str:
    value = _cell_with_aliases(row, resolved_columns, column)
    if value:
        return value
    return ""


def _select_sheet(workbook, sheet_aliases: list[str]):
    alias_keys = {_normalized_lookup_key(alias) for alias in sheet_aliases}
    for sheet_name in workbook.sheetnames:
        if _normalized_lookup_key(sheet_name) in alias_keys:
            return workbook[sheet_name]
    return workbook.active


def parse_items_import_workbook(file_obj: BinaryIO) -> ItemImportResult:
    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    try:
        sheet = _select_sheet(workbook, ITEM_SHEET_ALIASES)
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        headers = _header_map(header_row or [])
        resolved_columns = _resolve_columns(headers, list(ITEM_COLUMN_ALIASES), ITEM_COLUMN_ALIASES)

        rows: list[ItemImportRow] = []
        errors: list[ImportErrorDetail] = []

        for row_number, raw_row in enumerate(rows_iter, start=2):
            values = [_as_text(value) for value in raw_row]
            if not any(values):
                continue

            sku = _cell_with_aliases(values, resolved_columns, "Артикул")
            name = _cell_with_aliases(values, resolved_columns, "Наименование")
            unit_code = _cell_with_aliases(values, resolved_columns, "Единица")

            for column, message in REQUIRED_COLUMNS.items():
                if not _required_cell(values, resolved_columns, column):
                    errors.append(ImportErrorDetail(row_number=row_number, message=message))

            rows.append(
                ItemImportRow(
                    row_number=row_number,
                    sku=sku,
                    name=name,
                    unit_code=unit_code,
                    is_active=_as_bool(_cell_with_aliases(values, resolved_columns, "Активна")),
                    comment=_cell_with_aliases(values, resolved_columns, "Комментарий"),
                )
            )

        return ItemImportResult(rows=rows, errors=errors)
    finally:
        workbook.close()


def parse_opening_inventory_import_workbook(file_obj: BinaryIO) -> OpeningInventoryImportResult:
    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    try:
        sheet = _select_sheet(workbook, OPENING_INVENTORY_SHEET_ALIASES)
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        headers = _header_map(header_row or [])
        resolved_columns = _resolve_columns(
            headers,
            list(OPENING_INVENTORY_COLUMN_ALIASES),
            OPENING_INVENTORY_COLUMN_ALIASES,
        )

        rows: list[OpeningInventoryImportRow] = []
        errors: list[ImportErrorDetail] = []

        for row_number, raw_row in enumerate(rows_iter, start=2):
            values = [_as_text(value) for value in raw_row]
            if not any(values):
                continue

            warehouse_code = _cell_with_aliases(values, resolved_columns, "Склад")
            sku = _cell_with_aliases(values, resolved_columns, "Артикул")
            quantity_text = _cell_with_aliases(values, resolved_columns, "Фактическое количество")
            actual_quantity, quantity_ok = _as_decimal(quantity_text)

            for column, message in OPENING_INVENTORY_REQUIRED_COLUMNS.items():
                if not _required_cell(values, resolved_columns, column):
                    errors.append(ImportErrorDetail(row_number=row_number, message=message))

            if quantity_text and not quantity_ok:
                errors.append(ImportErrorDetail(row_number=row_number, message="Фактическое количество должно быть числом"))
            if quantity_ok and actual_quantity < 0:
                errors.append(ImportErrorDetail(row_number=row_number, message="Фактическое количество не может быть отрицательным"))

            rows.append(
                OpeningInventoryImportRow(
                    row_number=row_number,
                    warehouse_code=warehouse_code,
                    sku=sku,
                    actual_quantity=actual_quantity,
                    comment=_cell_with_aliases(values, resolved_columns, "Комментарий"),
                )
            )

        return OpeningInventoryImportResult(rows=rows, errors=errors)
    finally:
        workbook.close()


def _normalize_item_import_mode(import_mode: str) -> str:
    if import_mode in ITEM_IMPORT_MODES:
        return import_mode
    return ITEM_IMPORT_MODE_CREATE_ONLY


def validate_items_import_result(
    result: ItemImportResult,
    *,
    import_mode: str = ITEM_IMPORT_MODE_CREATE_ONLY,
    auto_create_units: bool = False,
) -> list[ImportErrorDetail]:
    import_mode = _normalize_item_import_mode(import_mode)
    errors = list(result.errors)
    seen_skus: set[str] = set()
    skus = {row.sku for row in result.rows if row.sku}
    unit_codes = {row.unit_code for row in result.rows if row.unit_code}
    existing_skus = set(Item.objects.filter(sku__in=skus).values_list("sku", flat=True))
    existing_unit_codes = set(Unit.objects.filter(code__in=unit_codes).values_list("code", flat=True))

    for row in result.rows:
        if row.sku:
            if row.sku in seen_skus:
                errors.append(ImportErrorDetail(row_number=row.row_number, message="Артикул повторяется в файле"))
            else:
                seen_skus.add(row.sku)
            if import_mode == ITEM_IMPORT_MODE_CREATE_ONLY and row.sku in existing_skus:
                errors.append(ImportErrorDetail(row_number=row.row_number, message="Артикул уже существует"))
            if import_mode == ITEM_IMPORT_MODE_UPDATE_EXISTING and row.sku not in existing_skus:
                errors.append(ImportErrorDetail(row_number=row.row_number, message="Артикул не найден для обновления"))

        if row.unit_code and row.unit_code not in existing_unit_codes:
            if not auto_create_units:
                errors.append(ImportErrorDetail(row_number=row.row_number, message="Единица не найдена"))
            elif len(row.unit_code) > Unit._meta.get_field("code").max_length:
                errors.append(
                    ImportErrorDetail(
                        row_number=row.row_number,
                        message="Код единицы измерения слишком длинный (максимум 20 символов)",
                    )
                )

    return errors


def validate_opening_inventory_import_result(result: OpeningInventoryImportResult) -> list[ImportErrorDetail]:
    errors = list(result.errors)
    seen_pairs: set[tuple[str, str]] = set()
    warehouse_codes = {row.warehouse_code for row in result.rows if row.warehouse_code}
    skus = {row.sku for row in result.rows if row.sku}
    existing_warehouse_codes = set(Warehouse.objects.filter(code__in=warehouse_codes).values_list("code", flat=True))
    existing_skus = set(Item.objects.filter(sku__in=skus).values_list("sku", flat=True))

    for row in result.rows:
        if row.warehouse_code and row.warehouse_code not in existing_warehouse_codes:
            errors.append(ImportErrorDetail(row_number=row.row_number, message="Склад не найден"))
        if row.sku and row.sku not in existing_skus:
            errors.append(ImportErrorDetail(row_number=row.row_number, message="Артикул не найден"))
        if row.warehouse_code and row.sku:
            pair = (row.warehouse_code, row.sku)
            if pair in seen_pairs:
                errors.append(ImportErrorDetail(row_number=row.row_number, message="Артикул повторяется для склада в файле"))
            else:
                seen_pairs.add(pair)

    return errors


def commit_items_import(
    result: ItemImportResult,
    *,
    import_mode: str = ITEM_IMPORT_MODE_CREATE_ONLY,
    auto_create_units: bool = False,
) -> ItemImportCommitResult:
    import_mode = _normalize_item_import_mode(import_mode)
    errors = validate_items_import_result(
        result,
        import_mode=import_mode,
        auto_create_units=auto_create_units,
    )
    if errors:
        return ItemImportCommitResult(created_count=0, updated_count=0, errors=errors)

    unit_codes = {row.unit_code for row in result.rows if row.unit_code}
    units = Unit.objects.in_bulk(unit_codes, field_name="code")
    missing_unit_codes = {unit_code for unit_code in unit_codes if unit_code not in units} if auto_create_units else set()

    with transaction.atomic():
        existing_items = {}
        if import_mode == ITEM_IMPORT_MODE_UPDATE_EXISTING:
            existing_items = Item.objects.in_bulk([row.sku for row in result.rows], field_name="sku")
            missing_item_errors = [
                ImportErrorDetail(row_number=row.row_number, message="Артикул не найден для обновления")
                for row in result.rows
                if row.sku not in existing_items
            ]
            if missing_item_errors:
                return ItemImportCommitResult(created_count=0, updated_count=0, errors=missing_item_errors)

        if missing_unit_codes:
            Unit.objects.bulk_create(
                [Unit(code=unit_code, name=unit_code) for unit_code in sorted(missing_unit_codes)],
                ignore_conflicts=True,
            )
            units = Unit.objects.in_bulk(unit_codes, field_name="code")

        if import_mode == ITEM_IMPORT_MODE_UPDATE_EXISTING:
            items_to_update = []
            race_errors: list[ImportErrorDetail] = []
            for row in result.rows:
                item = existing_items.get(row.sku)
                unit = units.get(row.unit_code)
                if unit is None:
                    race_errors.append(ImportErrorDetail(row_number=row.row_number, message="Единица не найдена"))
                    continue
                item.name = row.name
                item.unit = unit
                item.is_active = row.is_active
                item.notes = row.comment
                items_to_update.append(item)
            if race_errors:
                return ItemImportCommitResult(created_count=0, updated_count=0, errors=race_errors)
            Item.objects.bulk_update(items_to_update, ["name", "unit", "is_active", "notes"])

            return ItemImportCommitResult(created_count=0, updated_count=len(items_to_update), errors=[])

        items = []
        race_errors: list[ImportErrorDetail] = []
        for row in result.rows:
            unit = units.get(row.unit_code)
            if unit is None:
                race_errors.append(ImportErrorDetail(row_number=row.row_number, message="Единица не найдена"))
                continue
            items.append(
                Item(
                    sku=row.sku,
                    name=row.name,
                    unit=unit,
                    is_active=row.is_active,
                    notes=row.comment,
                )
            )
        if race_errors:
            return ItemImportCommitResult(created_count=0, updated_count=0, errors=race_errors)
        Item.objects.bulk_create(items)

    return ItemImportCommitResult(created_count=len(result.rows), updated_count=0, errors=[])


def commit_opening_inventory_import(result: OpeningInventoryImportResult) -> OpeningInventoryImportCommitResult:
    errors = validate_opening_inventory_import_result(result)
    warehouse_codes = {row.warehouse_code for row in result.rows if row.warehouse_code}
    if len(warehouse_codes) > 1:
        errors.append(ImportErrorDetail(row_number=0, message="Один импорт должен относиться к одному складу"))
    if errors:
        return OpeningInventoryImportCommitResult(inventory=None, created_lines_count=0, errors=errors)
    if not result.rows:
        return OpeningInventoryImportCommitResult(
            inventory=None,
            created_lines_count=0,
            errors=[ImportErrorDetail(row_number=0, message="Файл не содержит строк для импорта")],
        )

    warehouse = Warehouse.objects.get(code=result.rows[0].warehouse_code)
    items = Item.objects.in_bulk([row.sku for row in result.rows], field_name="sku")
    with transaction.atomic():
        inventory = InventoryDocument.objects.create(
            warehouse=warehouse,
            inventory_date=timezone.localdate(),
            scope=InventoryScope.FULL,
            comment="Импорт стартовых остатков из Excel. Проверьте строки перед проведением.",
        )
        InventoryLine.objects.bulk_create(
            [
                InventoryLine(
                    inventory=inventory,
                    item=items[row.sku],
                    actual_quantity=row.actual_quantity,
                    comment=row.comment,
                )
                for row in result.rows
            ]
        )

    return OpeningInventoryImportCommitResult(
        inventory=inventory,
        created_lines_count=len(result.rows),
        errors=[],
    )
