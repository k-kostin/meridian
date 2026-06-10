from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import sqlite3
import tempfile
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import DatabaseError, IntegrityError
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .backups import BackupError, configured_backup_paths, create_local_backup, create_pre_migration_backup_if_needed, restore_local_backup
from .demo import seed_demo_data
from .version import APP_VERSION_LABEL
from .models import (
    ActivityEvent,
    ActivityEventType,
    BackupKind,
    BackupRecord,
    BackupStatus,
    DocumentStatus,
    InventoryDocument,
    InventoryLine,
    InventoryScope,
    Item,
    ItemCategory,
    UserSavedView,
    StockDocument,
    StockDocumentLine,
    StockDocumentType,
    Unit,
    UserProfile,
    UserRole,
    Warehouse,
)
from .services import (
    PRESENTATION_BY_WAREHOUSE,
    PRESENTATION_CONSOLIDATED,
    build_daily_ledger,
    build_monthly_ledger,
    build_period_report,
    export_balances_xlsx,
    export_daily_ledger_xlsx,
    export_inventories_xlsx,
    export_monthly_ledger_xlsx,
    export_movements_xlsx,
    export_period_analysis_xlsx,
    get_balance_map,
    get_balance_rows,
    resolve_period,
)


class BackupRecordModelTests(TestCase):
    def test_backup_record_string_contains_kind_and_status(self):
        record = BackupRecord.objects.create(
            kind=BackupKind.MANUAL,
            status=BackupStatus.CREATED,
            backup_path="/tmp/meridian-20260610-120000-manual.sqlite3",
            source_database_path="/tmp/db.sqlite3",
            size_bytes=128,
            sha256="a" * 64,
            app_version="v0.5.0-dev",
        )

        self.assertIn("manual", str(record))
        self.assertIn("created", str(record))


class LocalBackupServiceTests(TestCase):
    def test_create_local_backup_copies_sqlite_database_and_records_metadata(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db_path = temp_path / "db.sqlite3"
            backup_dir = temp_path / "backups"

            with sqlite3.connect(db_path) as connection:
                connection.execute("CREATE TABLE sample (id INTEGER PRIMARY KEY, name TEXT)")
                connection.execute("INSERT INTO sample (name) VALUES ('alpha')")
                connection.commit()

            record = create_local_backup(
                database_path=db_path,
                backup_dir=backup_dir,
                kind=BackupKind.MANUAL,
                app_version="v0.5.0-dev",
                message="Manual backup created.",
            )

            self.assertEqual(record.status, BackupStatus.CREATED)
            self.assertEqual(record.kind, BackupKind.MANUAL)
            self.assertTrue(Path(record.backup_path).exists())
            self.assertGreater(record.size_bytes, 0)
            self.assertEqual(len(record.sha256), 64)

            with sqlite3.connect(record.backup_path) as backup_connection:
                rows = list(backup_connection.execute("SELECT name FROM sample"))

            self.assertEqual(rows, [("alpha",)])

    def test_create_local_backup_fails_for_missing_database(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)

            with self.assertRaises(BackupError):
                create_local_backup(
                    database_path=temp_path / "missing.sqlite3",
                    backup_dir=temp_path / "backups",
                    kind=BackupKind.MANUAL,
                    app_version="v0.5.0-dev",
                )

    def test_create_pre_migration_backup_skips_missing_database(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            with override_settings(
                DATABASES={"default": {"ENGINE": "django.db.backends.sqlite3", "NAME": temp_path / "db.sqlite3"}},
                WAREHOUSE_DATA_DIR=temp_path,
            ):
                record = create_pre_migration_backup_if_needed()

        self.assertIsNone(record)

    def test_configured_backup_paths_accepts_string_data_dir(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            with override_settings(
                DATABASES={"default": {"ENGINE": "django.db.backends.sqlite3", "NAME": temp_path / "db.sqlite3"}},
                WAREHOUSE_DATA_DIR=str(temp_path),
            ):
                paths = configured_backup_paths()

        self.assertEqual(paths.backup_dir, temp_path / "backups")

    def test_pre_migration_backup_survives_missing_metadata_table(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db_path = temp_path / "db.sqlite3"
            backup_dir = temp_path / "backups"

            with sqlite3.connect(db_path) as connection:
                connection.execute("CREATE TABLE sample (id INTEGER PRIMARY KEY)")
                connection.commit()

            with patch.object(BackupRecord, "save", side_effect=DatabaseError("no such table")):
                record = create_local_backup(
                    database_path=db_path,
                    backup_dir=backup_dir,
                    kind=BackupKind.PRE_MIGRATION,
                    app_version="v0.5.0-dev",
                    allow_unrecorded=True,
                )

            self.assertIsNone(record.pk)
            self.assertTrue(Path(record.backup_path).exists())
            self.assertEqual(record.kind, BackupKind.PRE_MIGRATION)

    def test_manual_backup_fails_when_metadata_cannot_be_saved(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db_path = temp_path / "db.sqlite3"

            with sqlite3.connect(db_path) as connection:
                connection.execute("CREATE TABLE sample (id INTEGER PRIMARY KEY)")
                connection.commit()

            with patch.object(BackupRecord, "save", side_effect=DatabaseError("metadata unavailable")):
                with self.assertRaises(BackupError):
                    create_local_backup(
                        database_path=db_path,
                        backup_dir=temp_path / "backups",
                        kind=BackupKind.MANUAL,
                        app_version="v0.5.0-dev",
                    )

    def test_restore_removes_target_wal_and_shm_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            backup_path = temp_path / "backup.sqlite3"
            target_path = temp_path / "db.sqlite3"

            with sqlite3.connect(backup_path) as connection:
                connection.execute("CREATE TABLE sample (name TEXT)")
                connection.execute("INSERT INTO sample (name) VALUES ('restored')")
                connection.commit()

            with sqlite3.connect(target_path) as connection:
                connection.execute("CREATE TABLE sample (name TEXT)")
                connection.execute("INSERT INTO sample (name) VALUES ('old')")
                connection.commit()

            wal_path = target_path.with_name(target_path.name + "-wal")
            shm_path = target_path.with_name(target_path.name + "-shm")
            wal_path.write_bytes(b"old wal")
            shm_path.write_bytes(b"old shm")

            restore_local_backup(backup_path=backup_path, database_path=target_path)

            self.assertFalse(wal_path.exists())
            self.assertFalse(shm_path.exists())
            with sqlite3.connect(target_path) as connection:
                rows = list(connection.execute("SELECT name FROM sample"))

            self.assertEqual(rows, [("restored",)])


class LocalBackupCommandTests(TestCase):
    def test_create_local_backup_command_creates_record(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db_path = temp_path / "db.sqlite3"

            with sqlite3.connect(db_path) as connection:
                connection.execute("CREATE TABLE sample (id INTEGER PRIMARY KEY)")
                connection.commit()

            with override_settings(
                DATABASES={"default": {"ENGINE": "django.db.backends.sqlite3", "NAME": db_path}},
                WAREHOUSE_DATA_DIR=temp_path,
            ):
                call_command("create_local_backup", verbosity=0)

            self.assertEqual(BackupRecord.objects.count(), 1)
            self.assertEqual(BackupRecord.objects.get().kind, BackupKind.MANUAL)

    def test_restore_local_backup_requires_confirm(self):
        with self.assertRaises(CommandError):
            call_command("restore_local_backup", "/tmp/example.sqlite3", verbosity=0)


class BackupViewTests(TestCase):
    def test_viewer_cannot_open_backup_list(self):
        user = User.objects.create_user(username="viewer", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.VIEWER)
        self.client.force_login(user)

        response = self.client.get(reverse("backup_list"))

        self.assertEqual(response.status_code, 403)

    def test_admin_can_create_backup_from_ui(self):
        user = User.objects.create_user(username="admin", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.ADMIN)
        self.client.force_login(user)

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db_path = temp_path / "db.sqlite3"

            with sqlite3.connect(db_path) as connection:
                connection.execute("CREATE TABLE sample (id INTEGER PRIMARY KEY)")
                connection.commit()

            with override_settings(
                DATABASES={"default": {"ENGINE": "django.db.backends.sqlite3", "NAME": db_path}},
                WAREHOUSE_DATA_DIR=temp_path,
            ):
                response = self.client.post(reverse("backup_create"), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(BackupRecord.objects.count(), 1)
        self.assertContains(response, "Резервная копия создана")

    def test_admin_backup_create_records_operational_activity(self):
        user = User.objects.create_user(username="admin-backup-audit", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.ADMIN)
        self.client.force_login(user)

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db_path = temp_path / "db.sqlite3"

            with sqlite3.connect(db_path) as connection:
                connection.execute("CREATE TABLE sample (id INTEGER PRIMARY KEY)")
                connection.commit()

            with override_settings(
                DATABASES={"default": {"ENGINE": "django.db.backends.sqlite3", "NAME": db_path}},
                WAREHOUSE_DATA_DIR=temp_path,
            ):
                response = self.client.post(reverse("backup_create"), follow=True)

        self.assertEqual(response.status_code, 200)
        event = ActivityEvent.objects.get(event_type=ActivityEventType.MANUAL_BACKUP_CREATED)
        self.assertEqual(event.actor, user)
        self.assertEqual(event.actor_label, "admin-backup-audit")
        self.assertEqual(event.metadata["backup_kind"], BackupKind.MANUAL)
        self.assertIn("backup_id", event.metadata)

    def test_admin_backup_create_succeeds_if_operational_activity_logging_fails(self):
        user = User.objects.create_user(username="admin-backup-resilient", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.ADMIN)
        self.client.force_login(user)

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db_path = temp_path / "db.sqlite3"

            with sqlite3.connect(db_path) as connection:
                connection.execute("CREATE TABLE sample (id INTEGER PRIMARY KEY)")
                connection.commit()

            with override_settings(
                DATABASES={"default": {"ENGINE": "django.db.backends.sqlite3", "NAME": db_path}},
                WAREHOUSE_DATA_DIR=temp_path,
            ):
                with patch("warehouse_app.views.record_manual_backup_created", side_effect=RuntimeError("audit down")):
                    with self.assertLogs("warehouse_app.views", level="ERROR") as logs:
                        response = self.client.post(reverse("backup_create"), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(BackupRecord.objects.count(), 1)
        self.assertContains(response, "Резервная копия создана")
        self.assertTrue(any("Failed to record manual backup activity event" in log for log in logs.output))


@override_settings(DEMO_MODE=True)
class WarehouseFlowTests(TestCase):
    def setUp(self):
        self.unit = Unit.objects.create(code="kg", name="Килограмм")
        self.warehouse = Warehouse.objects.create(code="main", name="Основной склад")
        self.item = Item.objects.create(sku="A-100", name="Позиция A", unit=self.unit)
        self.second_item = Item.objects.create(sku="B-200", name="Позиция B", unit=self.unit)

    def test_healthz_returns_readiness_payload(self):
        response = self.client.get("/healthz/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"status": "ok"})

    def test_authenticated_viewer_cannot_open_document_create(self):
        user = User.objects.create_user(username="viewer", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.VIEWER)
        self.client.force_login(user)

        response = self.client.get("/documents/new/")

        self.assertEqual(response.status_code, 403)

    def test_authenticated_operator_can_open_document_create(self):
        user = User.objects.create_user(username="operator", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.OPERATOR)
        self.client.force_login(user)

        response = self.client.get("/documents/new/")

        self.assertEqual(response.status_code, 200)

    @override_settings(DEMO_MODE=True)
    def test_anonymous_local_user_keeps_current_document_create_flow(self):
        response = self.client.get("/documents/new/")

        self.assertEqual(response.status_code, 200)

    @override_settings(DEMO_MODE=False)
    def test_anonymous_production_user_cannot_open_document_create(self):
        response = self.client.get("/documents/new/")

        self.assertEqual(response.status_code, 403)

    @override_settings(DEMO_MODE=False)
    def test_anonymous_production_user_sees_guest_role_label(self):
        response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Гость")
        self.assertNotContains(response, "Локальный режим")

    @override_settings(DEMO_MODE=True)
    def test_anonymous_demo_user_keeps_current_document_create_flow(self):
        response = self.client.get("/documents/new/")

        self.assertEqual(response.status_code, 200)

    def test_superuser_gets_admin_role(self):
        user = User.objects.create_superuser(username="root", password="pass")
        self.client.force_login(user)

        response = self.client.get("/items/import/")

        self.assertEqual(response.status_code, 200)

    def test_authenticated_user_without_profile_does_not_create_profile_on_read(self):
        user = User.objects.create_user(username="no-profile", password="pass")
        self.client.force_login(user)

        response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        self.assertFalse(UserProfile.objects.filter(user=user).exists())

    def test_viewer_cannot_post_document(self):
        document = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=timezone.localdate(),
        )
        StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("3"))
        user = User.objects.create_user(username="viewer-post", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.VIEWER)
        self.client.force_login(user)

        response = self.client.post(f"/documents/{document.pk}/post/")

        self.assertEqual(response.status_code, 403)

    def test_operator_cannot_edit_unit(self):
        user = User.objects.create_user(username="operator-unit", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.OPERATOR)
        self.client.force_login(user)

        response = self.client.get(f"/units/{self.unit.pk}/edit/")

        self.assertEqual(response.status_code, 403)

    def test_admin_can_edit_unit(self):
        user = User.objects.create_user(username="admin-unit", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.ADMIN)
        self.client.force_login(user)

        response = self.client.get(f"/units/{self.unit.pk}/edit/")

        self.assertEqual(response.status_code, 200)

    def test_viewer_can_read_reference_list(self):
        user = User.objects.create_user(username="viewer-ref", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.VIEWER)
        self.client.force_login(user)

        response = self.client.get("/units/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.unit.name)

    def test_viewer_sees_role_badge_and_no_document_create_cta(self):
        user = User.objects.create_user(username="viewer-ui", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.VIEWER)
        self.client.force_login(user)

        response = self.client.get("/documents/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Наблюдатель")
        self.assertNotContains(response, "Новый приход")
        self.assertContains(response, "Выгрузить Excel")

    def test_login_page_renders(self):
        response = self.client.get("/accounts/login/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Вход в систему")

    def test_item_category_can_be_assigned_to_item(self):
        category = ItemCategory.objects.create(name="Расходники", code="consumables")
        self.item.category = category
        self.item.save()

        self.item.refresh_from_db()
        self.assertEqual(self.item.category, category)

    def test_item_list_filters_by_category(self):
        category = ItemCategory.objects.create(name="Расходники", code="consumables")
        self.item.category = category
        self.item.save()
        other = Item.objects.create(sku="OTHER-CAT", name="Другая позиция", unit=self.unit)

        response = self.client.get("/items/", {"category": str(category.pk)})

        self.assertContains(response, self.item.sku)
        self.assertNotContains(response, other.sku)

    def test_parse_items_import_workbook_returns_rows(self):
        from .imports import parse_items_import_workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Номенклатура"
        sheet.append(["Артикул", "Наименование", "Единица", "Активна", "Комментарий"])
        sheet.append(["SKU-001", "Позиция импорта", "kg", "да", "тестовая строка"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = parse_items_import_workbook(buffer)

        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.errors, [])
        row = result.rows[0]
        self.assertEqual(row.row_number, 2)
        self.assertEqual(row.sku, "SKU-001")
        self.assertEqual(row.name, "Позиция импорта")
        self.assertEqual(row.unit_code, "kg")
        self.assertEqual(row.is_active, True)
        self.assertEqual(row.comment, "тестовая строка")

    def test_parse_items_import_workbook_reports_row_errors(self):
        from .imports import parse_items_import_workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Номенклатура"
        sheet.append(["Артикул", "Наименование", "Единица", "Активна", "Комментарий"])
        sheet.append(["", "", "", "нет", "неполная строка"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = parse_items_import_workbook(buffer)

        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.rows[0].row_number, 2)
        self.assertEqual(result.rows[0].is_active, False)
        self.assertEqual(
            [(error.row_number, error.message) for error in result.errors],
            [
                (2, "Артикул обязателен"),
                (2, "Наименование обязательно"),
                (2, "Единица обязательна"),
            ],
        )

    def test_parse_items_import_workbook_skips_empty_rows(self):
        from .imports import parse_items_import_workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Артикул", "Наименование", "Единица", "Активна", "Комментарий"])
        sheet.append([None, None, None, None, None])
        sheet.append(["SKU-002", "Вторая позиция", "pcs", "", ""])
        sheet.append(["", "", "", "", ""])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = parse_items_import_workbook(buffer)

        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.rows[0].sku, "SKU-002")
        self.assertEqual(result.rows[0].is_active, True)
        self.assertEqual(result.errors, [])

    def test_parse_items_import_workbook_accepts_case_insensitive_headers_and_common_false_values(self):
        from .imports import parse_items_import_workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["артикул", "НАИМЕНОВАНИЕ", "единица", "активна", "комментарий"])
        sheet.append(["SKU-003", "Третья позиция", "kg", "off", ""])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = parse_items_import_workbook(buffer)

        self.assertEqual(result.errors, [])
        self.assertEqual(result.rows[0].sku, "SKU-003")
        self.assertEqual(result.rows[0].is_active, False)

    def test_parse_items_import_workbook_uses_resolved_alias_column_consistently(self):
        from .imports import parse_items_import_workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Артикул", "Код", "Наименование", "Единица"])
        sheet.append(["", "FALLBACK-SKU", "Позиция с пустым артикулом", "pcs"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = parse_items_import_workbook(buffer)

        self.assertEqual(result.rows[0].sku, "")
        self.assertIn("Артикул обязателен", [error.message for error in result.errors])

    def _import_workbook_upload(self, rows, headers=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Номенклатура"
        sheet.append(headers or ["Артикул", "Наименование", "Единица", "Активна", "Комментарий"])
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "items.xlsx"
        return buffer

    def _opening_inventory_workbook_upload(self, rows, sheet_name="Стартовые остатки", headers=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(headers or ["Склад", "Артикул", "Фактическое количество", "Комментарий"])
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "opening-stock.xlsx"
        return buffer

    def test_parse_opening_inventory_import_workbook_returns_rows(self):
        from .imports import parse_opening_inventory_import_workbook

        buffer = self._opening_inventory_workbook_upload(
            [
                ["main", self.item.sku, 12.5, "остаток"],
            ]
        )

        result = parse_opening_inventory_import_workbook(buffer)

        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.rows[0].warehouse_code, "main")
        self.assertEqual(result.rows[0].sku, self.item.sku)
        self.assertEqual(result.rows[0].actual_quantity, Decimal("12.5"))
        self.assertEqual(result.rows[0].comment, "остаток")
        self.assertEqual(result.errors, [])

    def test_parse_opening_inventory_import_workbook_reports_required_fields(self):
        from .imports import parse_opening_inventory_import_workbook

        buffer = self._opening_inventory_workbook_upload([["", "", "", "bad"]])

        result = parse_opening_inventory_import_workbook(buffer)

        self.assertEqual(len(result.rows), 1)
        self.assertEqual(
            [error.message for error in result.errors],
            ["Склад обязателен", "Артикул обязателен", "Фактическое количество обязательно"],
        )

    def test_parse_opening_inventory_import_workbook_reports_invalid_quantity(self):
        from .imports import parse_opening_inventory_import_workbook

        buffer = self._opening_inventory_workbook_upload([["main", self.item.sku, "abc", "bad"]])

        result = parse_opening_inventory_import_workbook(buffer)

        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.rows[0].actual_quantity, Decimal("0"))
        self.assertEqual(result.errors[0].message, "Фактическое количество должно быть числом")

    def test_parse_opening_inventory_import_workbook_uses_resolved_alias_column_consistently(self):
        from .imports import parse_opening_inventory_import_workbook

        buffer = self._opening_inventory_workbook_upload(
            [[self.warehouse.code, "", "FALLBACK-SKU", "12", "остаток"]],
            headers=["Склад", "Артикул", "SKU", "Фактическое количество", "Комментарий"],
        )

        result = parse_opening_inventory_import_workbook(buffer)

        self.assertEqual(result.rows[0].sku, "")
        self.assertIn("Артикул обязателен", [error.message for error in result.errors])

    def test_parse_opening_inventory_import_workbook_accepts_realistic_sheet_and_headers(self):
        from .imports import parse_opening_inventory_import_workbook

        workbook = Workbook()
        workbook.active.title = "README"
        workbook.active.append(["Инструкция", "Не импортировать"])
        sheet = workbook.create_sheet("Остатки")
        sheet.append(["Код склада", "Код товара", "Кол-во", "Примечание"])
        sheet.append([self.warehouse.code, self.item.sku, "7,25", "остаток из файла"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = parse_opening_inventory_import_workbook(buffer)

        self.assertEqual(result.errors, [])
        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.rows[0].warehouse_code, self.warehouse.code)
        self.assertEqual(result.rows[0].sku, self.item.sku)
        self.assertEqual(result.rows[0].actual_quantity, Decimal("7.25"))
        self.assertEqual(result.rows[0].comment, "остаток из файла")

    def test_validate_opening_inventory_import_blocks_unknown_references_and_duplicates(self):
        from .imports import parse_opening_inventory_import_workbook, validate_opening_inventory_import_result

        buffer = self._opening_inventory_workbook_upload(
            [
                ["missing", self.item.sku, 1, ""],
                [self.warehouse.code, "missing-sku", 1, ""],
                [self.warehouse.code, self.item.sku, 1, ""],
                [self.warehouse.code, self.item.sku, 2, ""],
            ]
        )

        result = parse_opening_inventory_import_workbook(buffer)
        errors = validate_opening_inventory_import_result(result)

        self.assertIn("Склад не найден", [error.message for error in errors])
        self.assertIn("Артикул не найден", [error.message for error in errors])
        self.assertIn("Артикул повторяется для склада в файле", [error.message for error in errors])

    def test_commit_opening_inventory_import_creates_draft_full_inventory(self):
        from .imports import commit_opening_inventory_import, parse_opening_inventory_import_workbook

        buffer = self._opening_inventory_workbook_upload([[self.warehouse.code, self.item.sku, "5,5", "from file"]])
        result = parse_opening_inventory_import_workbook(buffer)

        commit_result = commit_opening_inventory_import(result)

        self.assertEqual(commit_result.errors, [])
        self.assertEqual(commit_result.created_lines_count, 1)
        inventory = commit_result.inventory
        self.assertIsNotNone(inventory)
        self.assertEqual(inventory.warehouse, self.warehouse)
        self.assertEqual(inventory.scope, InventoryScope.FULL)
        self.assertEqual(inventory.status, DocumentStatus.DRAFT)
        self.assertIn("Импорт стартовых остатков", inventory.comment)
        line = inventory.lines.get()
        self.assertEqual(line.item, self.item)
        self.assertEqual(line.actual_quantity, Decimal("5.5"))
        self.assertEqual(line.expected_quantity, Decimal("0"))
        self.assertEqual(line.comment, "from file")
        self.assertEqual(StockDocument.objects.count(), 0)

    def test_commit_opening_inventory_import_blocks_multiple_warehouses(self):
        from .imports import commit_opening_inventory_import, parse_opening_inventory_import_workbook

        other_warehouse = Warehouse.objects.create(code="reserve", name="Reserve")
        buffer = self._opening_inventory_workbook_upload(
            [
                [self.warehouse.code, self.item.sku, 5, ""],
                [other_warehouse.code, self.item.sku, 3, ""],
            ]
        )
        result = parse_opening_inventory_import_workbook(buffer)

        commit_result = commit_opening_inventory_import(result)

        self.assertEqual(commit_result.inventory, None)
        self.assertEqual(commit_result.created_lines_count, 0)
        self.assertIn("Один импорт должен относиться к одному складу", [error.message for error in commit_result.errors])
        self.assertEqual(InventoryDocument.objects.count(), 0)

    def test_item_import_preview_requires_reference_manager(self):
        viewer = User.objects.create_user(username="viewer-import", password="pass")
        UserProfile.objects.create(user=viewer, role=UserRole.VIEWER)
        self.client.force_login(viewer)

        viewer_response = self.client.get("/items/import/")

        self.assertEqual(viewer_response.status_code, 403)

        admin = User.objects.create_user(username="admin-import", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        admin_response = self.client.get("/items/import/")

        self.assertEqual(admin_response.status_code, 200)
        self.assertContains(admin_response, "Предпросмотр импорта номенклатуры")

    def test_item_import_preview_renders_valid_rows_without_creating_items(self):
        admin = User.objects.create_user(username="admin-import-preview", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)
        before_count = Item.objects.count()
        workbook = self._import_workbook_upload(
            [["SKU-IMPORT-1", "Импортная позиция", "kg", "да", "строка предпросмотра"]]
        )

        response = self.client.post("/items/import/", {"workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Item.objects.count(), before_count)
        self.assertContains(response, "SKU-IMPORT-1")
        self.assertContains(response, "Импортная позиция")
        self.assertContains(response, "строка предпросмотра")
        self.assertContains(response, "Импорт создает только новые позиции номенклатуры")

    def test_item_import_accepts_common_column_aliases(self):
        admin = User.objects.create_user(username="alias-admin", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        workbook = self._import_workbook_upload(
            [["ALIAS-001", "Позиция с алиасами", self.unit.code, "да", "алиас"]],
            headers=["SKU", "Название", "Ед.изм.", "Действует", "Примечание"],
        )

        response = self.client.post("/items/import/", {"workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "ALIAS-001")
        self.assertNotContains(response, "Единица обязательна")

    def test_parse_items_import_workbook_accepts_realistic_sheet_and_headers(self):
        from .imports import parse_items_import_workbook

        workbook = Workbook()
        workbook.active.title = "README"
        workbook.active.append(["Инструкция", "Не импортировать"])
        sheet = workbook.create_sheet("Товары")
        sheet.append(["Код товара", "Наименование товара", "Ед. изм.", "Активность", "Примечание"])
        sheet.append(["REAL-ITEM-1", "Реальная позиция", self.unit.code, "да", "живой шаблон"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = parse_items_import_workbook(buffer)

        self.assertEqual(result.errors, [])
        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.rows[0].sku, "REAL-ITEM-1")
        self.assertEqual(result.rows[0].name, "Реальная позиция")
        self.assertEqual(result.rows[0].unit_code, self.unit.code)
        self.assertEqual(result.rows[0].comment, "живой шаблон")

    def test_item_import_create_only_still_rejects_existing_sku(self):
        admin = User.objects.create_user(username="create-only-admin", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        workbook = self._import_workbook_upload([[self.item.sku, "Новое имя", self.unit.code, "да", ""]])

        response = self.client.post("/items/import/", {"workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Артикул уже существует")

    def test_item_import_update_mode_updates_existing_item(self):
        admin = User.objects.create_user(username="update-mode-admin", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        workbook = self._import_workbook_upload([[self.item.sku, "Обновленное имя", self.unit.code, "нет", "обновлено"]])

        response = self.client.post(
            "/items/import/",
            {"action": "commit", "import_mode": "update_existing", "workbook": workbook},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.item.refresh_from_db()
        self.assertEqual(self.item.name, "Обновленное имя")
        self.assertFalse(self.item.is_active)
        self.assertEqual(self.item.notes, "обновлено")
        self.assertContains(response, "Импорт обновил позиций: 1")

    def test_item_import_commit_records_operational_activity(self):
        admin = User.objects.create_user(username="admin-import-audit", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)
        workbook = self._import_workbook_upload(
            [["SKU-AUDIT-1", "Импорт с аудитом", self.unit.code, "да", ""]]
        )

        response = self.client.post("/items/import/", {"action": "commit", "workbook": workbook})

        self.assertRedirects(response, "/items/")
        event = ActivityEvent.objects.get(event_type=ActivityEventType.ITEM_IMPORT_COMMITTED)
        self.assertEqual(event.actor, admin)
        self.assertEqual(event.actor_label, "admin-import-audit")
        self.assertEqual(event.metadata["created_count"], 1)
        self.assertEqual(event.metadata["updated_count"], 0)
        self.assertEqual(event.metadata["import_mode"], "create_only")

    def test_item_import_update_mode_rejects_new_sku(self):
        admin = User.objects.create_user(username="update-mode-new-sku-admin", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        workbook = self._import_workbook_upload([["NEW-SKU", "Новая позиция", self.unit.code, "да", ""]])

        response = self.client.post(
            "/items/import/",
            {"action": "commit", "import_mode": "update_existing", "workbook": workbook},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Артикул не найден для обновления")
        self.assertFalse(Item.objects.filter(sku="NEW-SKU").exists())

    def test_item_import_rejects_unknown_unit_without_auto_create(self):
        admin = User.objects.create_user(username="unit-strict-admin", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        workbook = self._import_workbook_upload([["AUTO-UNIT-1", "Позиция", "box", "да", ""]])

        response = self.client.post("/items/import/", {"workbook": workbook})

        self.assertContains(response, "Единица не найдена")
        self.assertFalse(Unit.objects.filter(code="box").exists())

    def test_item_import_auto_creates_missing_unit_when_enabled(self):
        admin = User.objects.create_user(username="unit-auto-admin", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        workbook = self._import_workbook_upload([["AUTO-UNIT-2", "Позиция", "box", "да", ""]])

        response = self.client.post(
            "/items/import/",
            {"action": "commit", "import_mode": "create_only", "auto_create_units": "1", "workbook": workbook},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(Unit.objects.filter(code="box", name="box").exists())
        self.assertTrue(Item.objects.filter(sku="AUTO-UNIT-2", unit__code="box").exists())

    def test_item_import_auto_create_units_deduplicates_missing_unit_codes(self):
        admin = User.objects.create_user(username="unit-dedupe-admin", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        workbook = self._import_workbook_upload(
            [
                ["AUTO-UNIT-3", "Позиция 1", "pack", "да", ""],
                ["AUTO-UNIT-4", "Позиция 2", "pack", "да", ""],
            ]
        )

        response = self.client.post(
            "/items/import/",
            {"action": "commit", "import_mode": "create_only", "auto_create_units": "1", "workbook": workbook},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Unit.objects.filter(code="pack").count(), 1)
        self.assertEqual(Item.objects.filter(unit__code="pack").count(), 2)

    def test_item_import_auto_create_rejects_too_long_unit_code(self):
        admin = User.objects.create_user(username="unit-long-admin", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)
        long_unit_code = "x" * 21
        workbook = self._import_workbook_upload([["AUTO-UNIT-LONG", "Позиция", long_unit_code, "да", ""]])

        response = self.client.post(
            "/items/import/",
            {"action": "commit", "import_mode": "create_only", "auto_create_units": "1", "workbook": workbook},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Код единицы измерения слишком длинный")
        self.assertFalse(Unit.objects.filter(code=long_unit_code).exists())
        self.assertFalse(Item.objects.filter(sku="AUTO-UNIT-LONG").exists())

    def test_commit_items_import_create_mode_handles_missing_unit_after_validation(self):
        from .imports import commit_items_import, parse_items_import_workbook, validate_items_import_result

        temp_unit = Unit.objects.create(code="tmp", name="Временная единица")
        workbook = self._import_workbook_upload([["RACE-UNIT-1", "Позиция", temp_unit.code, "да", ""]])
        result = parse_items_import_workbook(workbook)

        self.assertEqual(validate_items_import_result(result), [])

        temp_unit.delete()
        commit_result = commit_items_import(result)

        self.assertEqual(commit_result.created_count, 0)
        self.assertEqual(commit_result.updated_count, 0)
        self.assertEqual([error.message for error in commit_result.errors], ["Единица не найдена"])
        self.assertFalse(Item.objects.filter(sku="RACE-UNIT-1").exists())

    def test_item_import_preview_renders_row_errors(self):
        admin = User.objects.create_user(username="admin-import-errors", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)
        workbook = self._import_workbook_upload([["", "", "", "нет", "ошибка строки"]])

        response = self.client.post("/items/import/", {"workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ошибки в строках")
        self.assertContains(response, "Артикул обязателен")
        self.assertContains(response, "Наименование обязательно")
        self.assertContains(response, "Единица обязательна")
        self.assertContains(response, "ошибка строки")

    def test_item_list_shows_import_cta_only_for_managers(self):
        viewer = User.objects.create_user(username="viewer-import-cta", password="pass")
        UserProfile.objects.create(user=viewer, role=UserRole.VIEWER)
        self.client.force_login(viewer)

        viewer_response = self.client.get("/items/")

        self.assertEqual(viewer_response.status_code, 200)
        self.assertNotContains(viewer_response, "Предпросмотр импорта")

        admin = User.objects.create_user(username="admin-import-cta", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        admin_response = self.client.get("/items/")

        self.assertEqual(admin_response.status_code, 200)
        self.assertContains(admin_response, "Предпросмотр импорта")

    def test_item_import_commit_creates_items_for_valid_workbook(self):
        admin = User.objects.create_user(username="admin-import-commit", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)
        workbook = self._import_workbook_upload(
            [["SKU-COMMIT-1", "Новая импортная позиция", "kg", "да", "создано импортом"]]
        )

        response = self.client.post("/items/import/", {"action": "commit", "workbook": workbook})

        self.assertRedirects(response, "/items/")
        item = Item.objects.get(sku="SKU-COMMIT-1")
        self.assertEqual(item.name, "Новая импортная позиция")
        self.assertEqual(item.unit, self.unit)
        self.assertEqual(item.is_active, True)
        self.assertEqual(item.notes, "создано импортом")

    def test_item_import_commit_blocks_unknown_unit(self):
        admin = User.objects.create_user(username="admin-import-unit", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)
        workbook = self._import_workbook_upload(
            [["SKU-COMMIT-2", "Позиция без единицы", "unknown", "да", ""]]
        )

        response = self.client.post("/items/import/", {"action": "commit", "workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Item.objects.filter(sku="SKU-COMMIT-2").exists())
        self.assertContains(response, "Единица не найдена")
        self.assertContains(response, "SKU-COMMIT-2")

    def test_item_import_commit_blocks_existing_sku(self):
        admin = User.objects.create_user(username="admin-import-duplicate", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)
        workbook = self._import_workbook_upload(
            [[self.item.sku, "Дубль существующей позиции", "kg", "да", ""]]
        )

        response = self.client.post("/items/import/", {"action": "commit", "workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Item.objects.filter(sku=self.item.sku).count(), 1)
        self.assertContains(response, "Артикул уже существует")

    def test_item_import_commit_blocks_duplicate_sku_in_workbook(self):
        admin = User.objects.create_user(username="admin-import-file-duplicate", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)
        workbook = self._import_workbook_upload(
            [
                ["SKU-DUP-FILE", "Первая строка", "kg", "да", ""],
                ["SKU-DUP-FILE", "Вторая строка", "kg", "да", ""],
            ]
        )

        response = self.client.post("/items/import/", {"action": "commit", "workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Item.objects.filter(sku="SKU-DUP-FILE").exists())
        self.assertContains(response, "Артикул повторяется в файле")

    def test_item_import_commit_blocks_parser_errors(self):
        admin = User.objects.create_user(username="admin-import-parser-errors", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)
        workbook = self._import_workbook_upload([["", "Без артикула", "kg", "да", ""]])

        response = self.client.post("/items/import/", {"action": "commit", "workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Артикул обязателен")
        self.assertEqual(Item.objects.filter(name="Без артикула").count(), 0)

    def test_viewer_cannot_commit_item_import(self):
        viewer = User.objects.create_user(username="viewer-import-commit", password="pass")
        UserProfile.objects.create(user=viewer, role=UserRole.VIEWER)
        self.client.force_login(viewer)
        workbook = self._import_workbook_upload(
            [["SKU-VIEWER-COMMIT", "Запрещенная позиция", "kg", "да", ""]]
        )

        response = self.client.post("/items/import/", {"action": "commit", "workbook": workbook})

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Item.objects.filter(sku="SKU-VIEWER-COMMIT").exists())

    def test_opening_inventory_import_preview_requires_stock_operator(self):
        viewer = User.objects.create_user(username="viewer-opening-import", password="pass")
        UserProfile.objects.create(user=viewer, role=UserRole.VIEWER)
        self.client.force_login(viewer)

        response = self.client.get("/inventories/import-opening/")

        self.assertEqual(response.status_code, 403)

    def test_inventory_list_shows_opening_import_cta_for_operator(self):
        operator = User.objects.create_user(username="operator-opening-import-cta", password="pass")
        UserProfile.objects.create(user=operator, role=UserRole.OPERATOR)
        self.client.force_login(operator)

        response = self.client.get("/inventories/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Импорт стартовых остатков")
        self.assertContains(response, "/inventories/import-opening/")

    def test_opening_inventory_import_preview_renders_valid_rows_without_creating_inventory(self):
        operator = User.objects.create_user(username="operator-opening-import", password="pass")
        UserProfile.objects.create(user=operator, role=UserRole.OPERATOR)
        self.client.force_login(operator)
        workbook = self._opening_inventory_workbook_upload([[self.warehouse.code, self.item.sku, 7, "preview"]])

        response = self.client.post("/inventories/import-opening/", {"workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Результат проверки")
        self.assertContains(response, self.item.sku)
        self.assertEqual(InventoryDocument.objects.count(), 0)

    def test_opening_inventory_import_accepts_common_column_aliases(self):
        operator = User.objects.create_user(username="opening-alias-operator", password="pass")
        UserProfile.objects.create(user=operator, role=UserRole.OPERATOR)
        self.client.force_login(operator)

        workbook = self._opening_inventory_workbook_upload(
            [[self.warehouse.code, self.item.sku, "12", "остаток"]],
            headers=["Код склада", "SKU", "Остаток", "Примечание"],
        )

        response = self.client.post("/inventories/import-opening/", {"workbook": workbook})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.item.sku)
        self.assertNotContains(response, "Фактическое количество обязательно")

    def test_opening_inventory_import_commit_redirects_to_draft_inventory(self):
        operator = User.objects.create_user(username="operator-opening-import-commit", password="pass")
        UserProfile.objects.create(user=operator, role=UserRole.OPERATOR)
        self.client.force_login(operator)
        workbook = self._opening_inventory_workbook_upload([[self.warehouse.code, self.item.sku, 7, "commit"]])

        response = self.client.post("/inventories/import-opening/", {"action": "commit", "workbook": workbook})

        inventory = InventoryDocument.objects.get()
        self.assertRedirects(response, f"/inventories/{inventory.pk}/")
        self.assertEqual(inventory.status, DocumentStatus.DRAFT)
        self.assertEqual(inventory.scope, InventoryScope.FULL)
        self.assertEqual(inventory.lines.count(), 1)

    def test_opening_inventory_import_commit_records_operational_activity(self):
        operator = User.objects.create_user(username="operator-opening-import-audit", password="pass")
        UserProfile.objects.create(user=operator, role=UserRole.OPERATOR)
        self.client.force_login(operator)
        workbook = self._opening_inventory_workbook_upload([[self.warehouse.code, self.item.sku, 7, "commit"]])

        response = self.client.post("/inventories/import-opening/", {"action": "commit", "workbook": workbook})

        inventory = InventoryDocument.objects.get()
        self.assertRedirects(response, f"/inventories/{inventory.pk}/")
        event = ActivityEvent.objects.get(event_type=ActivityEventType.OPENING_INVENTORY_IMPORT_COMMITTED)
        self.assertEqual(event.actor, operator)
        self.assertEqual(event.actor_label, "operator-opening-import-audit")
        self.assertEqual(event.inventory_document, inventory)
        self.assertEqual(event.warehouse, self.warehouse)
        self.assertEqual(event.metadata["created_lines_count"], 1)

    def _receipt(self, item, quantity):
        document = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=timezone.localdate(),
        )
        StockDocumentLine.objects.create(document=document, item=item, quantity=Decimal(quantity))
        document.post()
        return document

    def test_issue_cannot_go_below_zero(self):
        self._receipt(self.item, "10")
        issue = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=self.warehouse,
            operation_date=timezone.localdate(),
        )
        StockDocumentLine.objects.create(document=issue, item=self.item, quantity=Decimal("-12"))

        with self.assertRaises(ValidationError):
            issue.post()

        self.assertEqual(issue.status, DocumentStatus.DRAFT)

    def test_inventory_creates_adjustment_and_aligns_balance(self):
        self._receipt(self.item, "10")
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=timezone.localdate(),
            scope=InventoryScope.PARTIAL,
        )
        InventoryLine.objects.create(inventory=inventory, item=self.item, actual_quantity=Decimal("7"))

        inventory.post()

        balance = get_balance_map(self.warehouse)[self.item.id]
        adjustment = inventory.generated_documents.get()

        self.assertEqual(balance, Decimal("7"))
        self.assertEqual(inventory.status, DocumentStatus.POSTED)
        self.assertEqual(adjustment.document_type, StockDocumentType.ADJUSTMENT)
        self.assertEqual(adjustment.lines.get().quantity, Decimal("-3"))

    def test_full_inventory_zeroes_missing_items(self):
        self._receipt(self.item, "10")
        self._receipt(self.second_item, "5")
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=timezone.localdate(),
            scope=InventoryScope.FULL,
        )
        InventoryLine.objects.create(inventory=inventory, item=self.item, actual_quantity=Decimal("10"))

        inventory.post()

        balance = get_balance_map(self.warehouse)
        self.assertEqual(balance[self.item.id], Decimal("10"))
        self.assertNotIn(self.second_item.id, balance)

    def test_period_report_shows_opening_movements_and_closing(self):
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("10"))
        receipt.post()

        issue = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 5),
        )
        StockDocumentLine.objects.create(document=issue, item=self.item, quantity=Decimal("-3"))
        issue.post()

        second_receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 10),
        )
        StockDocumentLine.objects.create(document=second_receipt, item=self.item, quantity=Decimal("2"))
        second_receipt.post()

        report = build_period_report(
            warehouse=self.warehouse,
            period_start=date(2026, 3, 5),
            period_end=date(2026, 3, 10),
        )

        row = report["rows"][0]
        self.assertEqual(row["opening"], Decimal("10"))
        self.assertEqual(row["incoming"], Decimal("2"))
        self.assertEqual(row["outgoing"], Decimal("3"))
        self.assertEqual(row["net"], Decimal("-1"))
        self.assertEqual(row["closing"], Decimal("9"))

    def test_post_is_idempotent_for_document(self):
        document = self._receipt(self.item, "10")
        first_posted_at = document.posted_at

        document.post()
        document.refresh_from_db()

        self.assertEqual(document.status, DocumentStatus.POSTED)
        self.assertEqual(document.posted_at, first_posted_at)

    def test_posting_stock_document_records_activity_event_once(self):
        document = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 11),
            comment="timeline receipt",
        )
        StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("5"))

        document.post()
        document.post()

        events = ActivityEvent.objects.filter(stock_document=document)
        self.assertEqual(events.count(), 1)
        event = events.get()
        self.assertEqual(event.event_type, ActivityEventType.STOCK_DOCUMENT_POSTED)
        self.assertEqual(event.warehouse, self.warehouse)
        self.assertEqual(event.inventory_document, None)
        self.assertIn(document.number, event.message)
        self.assertIn("Приход", event.message)

    def test_posting_inventory_records_inventory_and_adjustment_activity(self):
        self._receipt(self.item, "10")
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 3, 12),
            scope=InventoryScope.PARTIAL,
            comment="timeline inventory",
        )
        InventoryLine.objects.create(
            inventory=inventory,
            item=self.item,
            actual_quantity=Decimal("7"),
        )

        inventory.post()
        inventory.post()

        adjustment = inventory.generated_documents.get()
        inventory_events = ActivityEvent.objects.filter(
            inventory_document=inventory,
            event_type__in=[
                ActivityEventType.INVENTORY_POSTED,
                ActivityEventType.INVENTORY_ADJUSTMENT_CREATED,
            ],
        )

        self.assertEqual(
            set(inventory_events.values_list("event_type", flat=True)),
            {
                ActivityEventType.INVENTORY_POSTED,
                ActivityEventType.INVENTORY_ADJUSTMENT_CREATED,
            },
        )
        self.assertEqual(inventory_events.count(), 2)
        adjustment_posted_event = ActivityEvent.objects.get(
            stock_document=adjustment,
            event_type=ActivityEventType.STOCK_DOCUMENT_POSTED,
        )
        self.assertEqual(adjustment_posted_event.inventory_document, inventory)

    def test_document_numbers_increment_for_same_day(self):
        first = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        second = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )

        self.assertEqual(first.number, "RCV-20260301-001")
        self.assertEqual(second.number, "RCV-20260301-002")

    def test_inventory_numbers_increment_for_same_day(self):
        first = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 3, 1),
            scope=InventoryScope.PARTIAL,
        )
        second = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 3, 1),
            scope=InventoryScope.PARTIAL,
        )

        self.assertEqual(first.number, "INV-20260301-001")
        self.assertEqual(second.number, "INV-20260301-002")

    def test_empty_inventory_cannot_be_posted(self):
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=timezone.localdate(),
            scope=InventoryScope.PARTIAL,
        )

        with self.assertRaises(ValidationError):
            inventory.post()

    def test_inventory_without_delta_does_not_create_adjustment(self):
        self._receipt(self.item, "10")
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=timezone.localdate(),
            scope=InventoryScope.PARTIAL,
        )
        InventoryLine.objects.create(inventory=inventory, item=self.item, actual_quantity=Decimal("10"))

        inventory.post()

        self.assertEqual(inventory.generated_documents.count(), 0)

    def test_resolve_period_month_and_custom(self):
        month_period = resolve_period(mode="month", anchor_date=date(2026, 3, 15))
        custom_period = resolve_period(mode="custom", date_from=date(2026, 3, 20), date_to=date(2026, 3, 10))

        self.assertEqual(month_period["start"], date(2026, 3, 1))
        self.assertEqual(month_period["end"], date(2026, 3, 31))
        self.assertEqual(custom_period["start"], date(2026, 3, 10))
        self.assertEqual(custom_period["end"], date(2026, 3, 20))

    def test_period_report_aggregates_same_item_across_warehouses(self):
        second_warehouse = Warehouse.objects.create(code="reserve", name="Резервный склад")

        receipt_main = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 2, 25),
        )
        StockDocumentLine.objects.create(document=receipt_main, item=self.item, quantity=Decimal("53"))
        receipt_main.post()

        receipt_reserve = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=second_warehouse,
            operation_date=date(2026, 3, 6),
        )
        StockDocumentLine.objects.create(document=receipt_reserve, item=self.item, quantity=Decimal("6"))
        receipt_reserve.post()

        issue_main = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 8),
        )
        StockDocumentLine.objects.create(document=issue_main, item=self.item, quantity=Decimal("-9"))
        issue_main.post()

        report = build_period_report(period_start=date(2026, 3, 1), period_end=date(2026, 3, 31))

        row = next(result for result in report["rows"] if result["item_id"] == self.item.id)
        self.assertEqual(row["opening"], Decimal("53"))
        self.assertEqual(row["incoming"], Decimal("6"))
        self.assertEqual(row["outgoing"], Decimal("9"))
        self.assertEqual(row["net"], Decimal("-3"))
        self.assertEqual(row["closing"], Decimal("50"))

    def test_period_report_can_split_rows_by_warehouse(self):
        second_warehouse = Warehouse.objects.create(code="reserve", name="Резервный склад")

        receipt_main = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt_main, item=self.item, quantity=Decimal("10"))
        receipt_main.post()

        receipt_reserve = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=second_warehouse,
            operation_date=date(2026, 3, 2),
        )
        StockDocumentLine.objects.create(document=receipt_reserve, item=self.item, quantity=Decimal("5"))
        receipt_reserve.post()

        issue_reserve = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=second_warehouse,
            operation_date=date(2026, 3, 5),
        )
        StockDocumentLine.objects.create(document=issue_reserve, item=self.item, quantity=Decimal("-2"))
        issue_reserve.post()

        report = build_period_report(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            presentation=PRESENTATION_BY_WAREHOUSE,
        )

        self.assertEqual(report["summary"]["rows_count"], 2)
        main_row = next(row for row in report["rows"] if row["warehouse_name"] == "Основной склад")
        reserve_row = next(row for row in report["rows"] if row["warehouse_name"] == "Резервный склад")

        self.assertEqual(main_row["incoming"], Decimal("10"))
        self.assertEqual(main_row["outgoing"], Decimal("0"))
        self.assertEqual(main_row["closing"], Decimal("10"))
        self.assertEqual(reserve_row["incoming"], Decimal("5"))
        self.assertEqual(reserve_row["outgoing"], Decimal("2"))
        self.assertEqual(reserve_row["closing"], Decimal("3"))

    def test_balance_rows_can_include_zero_positions_with_totals(self):
        zero_item = Item.objects.create(sku="ZZZ-EMPTY", name="Пустая позиция", unit=self.unit, is_active=False)

        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("10"))
        receipt.post()

        issue = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 2),
        )
        StockDocumentLine.objects.create(document=issue, item=self.item, quantity=Decimal("-4"))
        issue.post()

        rows = get_balance_rows(
            warehouse=self.warehouse,
            presentation=PRESENTATION_CONSOLIDATED,
            include_zero=True,
        )

        item_row = next(row for row in rows if row["item__sku"] == "A-100")
        zero_row = next(row for row in rows if row["item__sku"] == zero_item.sku)

        self.assertEqual(item_row["incoming_total"], Decimal("10"))
        self.assertEqual(item_row["outgoing_total"], Decimal("4"))
        self.assertEqual(item_row["quantity"], Decimal("6"))
        self.assertEqual(zero_row["incoming_total"], Decimal("0"))
        self.assertEqual(zero_row["outgoing_total"], Decimal("0"))
        self.assertEqual(zero_row["quantity"], Decimal("0"))

    def test_daily_ledger_matches_excel_light_shape(self):
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("10"))
        receipt.post()

        issue = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 3),
        )
        StockDocumentLine.objects.create(document=issue, item=self.item, quantity=Decimal("-4"))
        issue.post()

        ledger = build_daily_ledger(
            warehouse=self.warehouse,
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 3),
            presentation=PRESENTATION_CONSOLIDATED,
        )

        self.assertEqual(ledger["period"]["label"], "Дни за 03.2026")
        rows = [row for row in ledger["rows"] if row["sku"] == "A-100"]
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0]["date"], date(2026, 3, 1))
        self.assertEqual(rows[0]["opening"], Decimal("0"))
        self.assertEqual(rows[0]["incoming"], Decimal("10"))
        self.assertEqual(rows[0]["closing"], Decimal("10"))
        self.assertEqual(rows[1]["date"], date(2026, 3, 2))
        self.assertEqual(rows[1]["opening"], Decimal("10"))
        self.assertEqual(rows[1]["incoming"], Decimal("0"))
        self.assertEqual(rows[1]["closing"], Decimal("10"))
        self.assertEqual(rows[2]["date"], date(2026, 3, 3))
        self.assertEqual(rows[2]["opening"], Decimal("10"))
        self.assertEqual(rows[2]["outgoing"], Decimal("4"))
        self.assertEqual(rows[2]["net"], Decimal("-4"))
        self.assertEqual(rows[2]["closing"], Decimal("6"))

    def test_monthly_ledger_matches_excel_light_shape(self):
        feb_receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 2, 10),
        )
        StockDocumentLine.objects.create(document=feb_receipt, item=self.item, quantity=Decimal("8"))
        feb_receipt.post()

        march_issue = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 5),
        )
        StockDocumentLine.objects.create(document=march_issue, item=self.item, quantity=Decimal("-3"))
        march_issue.post()

        ledger = build_monthly_ledger(
            warehouse=self.warehouse,
            period_start=date(2026, 2, 1),
            period_end=date(2026, 3, 31),
            presentation=PRESENTATION_CONSOLIDATED,
        )

        rows = [row for row in ledger["rows"] if row["sku"] == "A-100"]
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["month_label"], "02.2026")
        self.assertEqual(rows[0]["opening"], Decimal("0"))
        self.assertEqual(rows[0]["incoming"], Decimal("8"))
        self.assertEqual(rows[0]["closing"], Decimal("8"))
        self.assertEqual(rows[1]["month_label"], "03.2026")
        self.assertEqual(rows[1]["opening"], Decimal("8"))
        self.assertEqual(rows[1]["outgoing"], Decimal("3"))
        self.assertEqual(rows[1]["closing"], Decimal("5"))

    def test_export_workbooks_include_metadata_and_unit_formats(self):
        pcs_unit = Unit.objects.create(code="pcs", name="Штука", display_precision=0)
        pcs_item = Item.objects.create(sku="C-300", name="Позиция C", unit=pcs_unit)

        first_receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=first_receipt, item=self.item, quantity=Decimal("10.125"))
        first_receipt.post()

        second_receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 2),
        )
        StockDocumentLine.objects.create(document=second_receipt, item=pcs_item, quantity=Decimal("7"))
        second_receipt.post()

        balances_export = export_balances_xlsx(presentation=PRESENTATION_CONSOLIDATED)
        balances_workbook = load_workbook(balances_export.buffer)
        self.assertEqual(balances_workbook.sheetnames[0], "Параметры")

        metadata_sheet = balances_workbook["Параметры"]
        metadata = {
            metadata_sheet.cell(row=row_number, column=1).value: metadata_sheet.cell(row=row_number, column=2).value
            for row_number in range(2, metadata_sheet.max_row + 1)
        }
        self.assertEqual(metadata["Представление"], "Сводно по складам")
        self.assertEqual(metadata["Показывать нулевые позиции"], "Нет")

        balance_sheet = balances_workbook["Остатки"]
        formats = {
            balance_sheet.cell(row=row_number, column=1).value: (
                balance_sheet.cell(row=row_number, column=4).number_format,
                balance_sheet.cell(row=row_number, column=5).number_format,
                balance_sheet.cell(row=row_number, column=6).number_format,
            )
            for row_number in range(2, balance_sheet.max_row + 1)
        }
        self.assertEqual(formats["A-100"], ("0.000", "0.000", "0.000"))
        self.assertEqual(formats["C-300"], ("0", "0", "0"))

        analysis_export = export_period_analysis_xlsx(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            label="Месяц 03.2026",
            presentation=PRESENTATION_BY_WAREHOUSE,
            mode_label="Месяц",
        )
        analysis_workbook = load_workbook(analysis_export.buffer)
        analysis_metadata_sheet = analysis_workbook["Параметры"]
        analysis_metadata = {
            analysis_metadata_sheet.cell(row=row_number, column=1).value: analysis_metadata_sheet.cell(
                row=row_number, column=2
            ).value
            for row_number in range(2, analysis_metadata_sheet.max_row + 1)
        }
        self.assertEqual(analysis_metadata["Режим периода"], "Месяц")
        self.assertEqual(analysis_metadata["Представление"], "С разбивкой по складам")

        movements_export = export_movements_xlsx()
        movements_workbook = load_workbook(movements_export.buffer)
        movements_sheet = movements_workbook["Движения"]
        self.assertFalse(isinstance(movements_sheet.cell(row=2, column=1).value, str))
        self.assertEqual(movements_sheet.cell(row=2, column=1).number_format, "DD.MM.YYYY")

        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 3, 3),
            scope=InventoryScope.PARTIAL,
        )
        InventoryLine.objects.create(
            inventory=inventory,
            item=self.item,
            actual_quantity=Decimal("10.125"),
            expected_quantity=Decimal("9.125"),
        )
        inventories_export = export_inventories_xlsx()
        inventories_workbook = load_workbook(inventories_export.buffer)
        inventories_sheet = inventories_workbook["Инвентаризации"]
        self.assertFalse(isinstance(inventories_sheet.cell(row=2, column=1).value, str))
        self.assertEqual(inventories_sheet.cell(row=2, column=1).number_format, "DD.MM.YYYY")

        daily_export = export_daily_ledger_xlsx(
            warehouse=self.warehouse,
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
        )
        daily_workbook = load_workbook(daily_export.buffer)
        daily_sheet = daily_workbook["Дни"]
        self.assertEqual(daily_sheet.cell(row=2, column=1).number_format, "DD.MM.YYYY")

        monthly_export = export_monthly_ledger_xlsx(
            warehouse=self.warehouse,
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
        )
        monthly_workbook = load_workbook(monthly_export.buffer)
        monthly_sheet = monthly_workbook["Месяцы"]
        self.assertEqual(monthly_sheet.cell(row=2, column=1).number_format, "MM.YYYY")

    def test_period_report_grand_total_consolidated(self):
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("10"))
        receipt.post()

        receipt2 = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 2),
        )
        StockDocumentLine.objects.create(document=receipt2, item=self.second_item, quantity=Decimal("5"))
        receipt2.post()

        report = build_period_report(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            presentation=PRESENTATION_CONSOLIDATED,
        )

        self.assertIsNone(report["grouped_rows"])
        # Both items share the same unit ("kg"), so one totals entry
        self.assertEqual(len(report["grand_total_by_unit"]), 1)
        gt = report["grand_total_by_unit"][0]
        self.assertEqual(gt["unit"], "kg")
        self.assertEqual(gt["incoming"], Decimal("15"))
        self.assertEqual(gt["outgoing"], Decimal("0"))
        self.assertEqual(gt["closing"], Decimal("15"))

    def test_balance_warehouse_label_hides_code_when_names_are_unique(self):
        self._receipt(self.item, "10")
        self._receipt(self.second_item, "5")

        rows = get_balance_rows(presentation=PRESENTATION_BY_WAREHOUSE)

        self.assertGreaterEqual(len(rows), 2)
        self.assertEqual({row["warehouse_label"] for row in rows}, {"Основной склад"})

    def test_period_report_totals_by_warehouse(self):
        second_warehouse = Warehouse.objects.create(code="reserve", name="Резервный склад")

        receipt_main = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt_main, item=self.item, quantity=Decimal("10"))
        receipt_main.post()

        receipt_reserve = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=second_warehouse,
            operation_date=date(2026, 3, 2),
        )
        StockDocumentLine.objects.create(document=receipt_reserve, item=self.item, quantity=Decimal("6"))
        receipt_reserve.post()

        issue_reserve = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=second_warehouse,
            operation_date=date(2026, 3, 3),
        )
        StockDocumentLine.objects.create(document=issue_reserve, item=self.item, quantity=Decimal("-2"))
        issue_reserve.post()

        report = build_period_report(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            presentation=PRESENTATION_BY_WAREHOUSE,
        )

        self.assertIsNotNone(report["grouped_rows"])
        self.assertEqual(len(report["grouped_rows"]), 2)

        main_group = next(g for g in report["grouped_rows"] if g["warehouse_name"] == "Основной склад")
        reserve_group = next(g for g in report["grouped_rows"] if g["warehouse_name"] == "Резервный склад")

        # Both warehouses stock only kg items, so one subtotal entry per warehouse
        self.assertEqual(len(main_group["subtotals_by_unit"]), 1)
        main_st = main_group["subtotals_by_unit"][0]
        self.assertEqual(main_st["unit"], "kg")
        self.assertEqual(main_st["incoming"], Decimal("10"))
        self.assertEqual(main_st["outgoing"], Decimal("0"))
        self.assertEqual(main_st["closing"], Decimal("10"))

        self.assertEqual(len(reserve_group["subtotals_by_unit"]), 1)
        reserve_st = reserve_group["subtotals_by_unit"][0]
        self.assertEqual(reserve_st["unit"], "kg")
        self.assertEqual(reserve_st["incoming"], Decimal("6"))
        self.assertEqual(reserve_st["outgoing"], Decimal("2"))
        self.assertEqual(reserve_st["closing"], Decimal("4"))

        # Grand total: single kg entry, all warehouses combined
        self.assertEqual(len(report["grand_total_by_unit"]), 1)
        gt = report["grand_total_by_unit"][0]
        self.assertEqual(gt["unit"], "kg")
        self.assertEqual(gt["incoming"], Decimal("16"))
        self.assertEqual(gt["outgoing"], Decimal("2"))
        self.assertEqual(gt["closing"], Decimal("14"))

    def test_totals_never_mix_different_units(self):
        """Grand total must produce separate entries for each unit, not one mixed number."""
        pcs_unit = Unit.objects.create(code="pcs", name="Штука", display_precision=0)
        pcs_item = Item.objects.create(sku="D-400", name="Позиция D", unit=pcs_unit)

        receipt_kg = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt_kg, item=self.item, quantity=Decimal("10"))
        receipt_kg.post()

        receipt_pcs = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 2),
        )
        StockDocumentLine.objects.create(document=receipt_pcs, item=pcs_item, quantity=Decimal("5"))
        receipt_pcs.post()

        report = build_period_report(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            presentation=PRESENTATION_CONSOLIDATED,
        )

        # Must be two separate unit entries — never a single mixed number
        self.assertEqual(len(report["grand_total_by_unit"]), 2)
        units = {t["unit"] for t in report["grand_total_by_unit"]}
        self.assertEqual(units, {"kg", "pcs"})

        kg_total = next(t for t in report["grand_total_by_unit"] if t["unit"] == "kg")
        pcs_total = next(t for t in report["grand_total_by_unit"] if t["unit"] == "pcs")

        self.assertEqual(kg_total["incoming"], Decimal("10"))
        self.assertEqual(pcs_total["incoming"], Decimal("5"))

        # Also verify by_warehouse presentation: subtotals must also split by unit
        report_bw = build_period_report(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            presentation=PRESENTATION_BY_WAREHOUSE,
        )
        main_group = report_bw["grouped_rows"][0]
        subtotal_units = {st["unit"] for st in main_group["subtotals_by_unit"]}
        self.assertEqual(subtotal_units, {"kg", "pcs"})

    def test_grouping_stable_by_warehouse_id_not_name(self):
        """Two warehouses with the same name must remain separate groups."""
        # Warehouse.name has no unique constraint — only code is unique
        duplicate_name_warehouse = Warehouse.objects.create(
            code="secondary", name=self.warehouse.name  # same name, different code/PK
        )

        receipt_main = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt_main, item=self.item, quantity=Decimal("10"))
        receipt_main.post()

        receipt_dup = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=duplicate_name_warehouse,
            operation_date=date(2026, 3, 2),
        )
        StockDocumentLine.objects.create(document=receipt_dup, item=self.item, quantity=Decimal("5"))
        receipt_dup.post()

        report = build_period_report(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            presentation=PRESENTATION_BY_WAREHOUSE,
        )

        # Must be 2 groups keyed by warehouse_id, not collapsed by name
        self.assertEqual(len(report["grouped_rows"]), 2)
        group_ids = {g["warehouse_id"] for g in report["grouped_rows"]}
        self.assertEqual(group_ids, {self.warehouse.pk, duplicate_name_warehouse.pk})

        # Group-level labels must be distinct and include the code
        labels = {g["warehouse_label"] for g in report["grouped_rows"]}
        self.assertEqual(len(labels), 2)
        for group in report["grouped_rows"]:
            self.assertIn(group["warehouse_name"], group["warehouse_label"])
            self.assertIn(group["warehouse_code"], group["warehouse_label"])

        # Row-level warehouse_label must also be distinct (used in template and Excel)
        row_labels = {row["warehouse_label"] for row in report["rows"]}
        self.assertEqual(len(row_labels), 2)
        for row in report["rows"]:
            self.assertIn(row["warehouse_code"], row["warehouse_label"])

        # Main Excel sheet must use the distinct label in the Склад column (col 1)
        export_payload = export_period_analysis_xlsx(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            label="Месяц 03.2026",
            presentation=PRESENTATION_BY_WAREHOUSE,
            mode_label="Месяц",
        )
        workbook = load_workbook(export_payload.buffer)
        sheet = workbook["Аналитика"]
        # Rows 2+ are data rows; column 1 is the warehouse label
        excel_warehouse_labels = {
            sheet.cell(row=r, column=1).value
            for r in range(2, sheet.max_row + 1)
        }
        self.assertEqual(len(excel_warehouse_labels), 2)
        for label_val in excel_warehouse_labels:
            self.assertIsNotNone(label_val)
            # Each label must contain the warehouse name and distinguish by code
            self.assertIn("(", label_val)  # code is in parentheses

    def test_export_analysis_has_totals_sheet(self):
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("10"))
        receipt.post()

        export_payload = export_period_analysis_xlsx(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            label="Месяц 03.2026",
            presentation=PRESENTATION_BY_WAREHOUSE,
            mode_label="Месяц",
        )
        workbook = load_workbook(export_payload.buffer)
        self.assertIn("Итоги", workbook.sheetnames)

        totals_sheet = workbook["Итоги"]
        # Row 1: header; Row 2: warehouse subtotal (kg); Row 3: grand total (kg)
        self.assertEqual(totals_sheet.max_row, 3)

        # by_warehouse columns: Склад(1), Ед.(2), На начало(3), Приход(4), Расход(5), Дельта(6), На конец(7)
        self.assertEqual(totals_sheet.cell(row=2, column=1).value, "Основной склад")
        self.assertEqual(totals_sheet.cell(row=2, column=2).value, "kg")
        self.assertEqual(totals_sheet.cell(row=2, column=4).value, Decimal("10"))  # incoming
        self.assertEqual(totals_sheet.cell(row=2, column=7).value, Decimal("10"))  # closing

        self.assertEqual(totals_sheet.cell(row=3, column=1).value, "Итого по всем складам")
        self.assertEqual(totals_sheet.cell(row=3, column=2).value, "kg")
        self.assertEqual(totals_sheet.cell(row=3, column=7).value, Decimal("10"))  # closing

        analysis_metadata_sheet = workbook["Параметры"]
        metadata = {
            analysis_metadata_sheet.cell(row=r, column=1).value: analysis_metadata_sheet.cell(row=r, column=2).value
            for r in range(2, analysis_metadata_sheet.max_row + 1)
        }
        self.assertIn("Лист «Итоги»", metadata)
        self.assertEqual(metadata["Доп. листы по складам"], "Да")

    def test_balance_report_supports_search_and_pagination(self):
        for index in range(12):
            item = Item.objects.create(sku=f"T-{index:03d}", name=f"Тестовая позиция {index}", unit=self.unit)
            receipt = StockDocument.objects.create(
                document_type=StockDocumentType.RECEIPT,
                warehouse=self.warehouse,
                operation_date=date(2026, 3, 1),
            )
            StockDocumentLine.objects.create(document=receipt, item=item, quantity=Decimal("1"))
            receipt.post()

        search_item = Item.objects.create(sku="SEARCH-220", name="Нужная позиция", unit=self.unit)
        search_receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 2),
        )
        StockDocumentLine.objects.create(document=search_receipt, item=search_item, quantity=Decimal("5"))
        search_receipt.post()

        paged_response = self.client.get("/balances/", {"presentation": "consolidated", "page_size": 10, "page": 2})
        self.assertEqual(paged_response.status_code, 200)
        self.assertEqual(paged_response.context["page_obj"].paginator.per_page, 10)
        self.assertEqual(paged_response.context["page_obj"].number, 2)

        filtered_response = self.client.get("/balances/", {"presentation": "consolidated", "q": "220"})
        self.assertEqual(filtered_response.status_code, 200)
        rows = list(filtered_response.context["balances"].object_list)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["item__sku"], "SEARCH-220")

    def test_balance_report_can_show_zero_positions(self):
        Item.objects.create(sku="ZERO-001", name="Нулевая позиция", unit=self.unit, is_active=False)
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("5"))
        receipt.post()

        response = self.client.get("/balances/", {"presentation": "consolidated", "include_zero": "1"})

        self.assertEqual(response.status_code, 200)
        rows = list(response.context["balances"].object_list)
        self.assertTrue(any(row["item__sku"] == "ZERO-001" for row in rows))

    def test_balances_can_filter_by_category(self):
        category = ItemCategory.objects.create(name="Расходники", code="consumables")
        self.item.category = category
        self.item.save()
        other = Item.objects.create(sku="NO-CAT-BAL", name="Без категории", unit=self.unit)
        self._receipt(self.item, "5")
        self._receipt(other, "7")

        response = self.client.get("/balances/", {"category": str(category.pk), "presentation": "consolidated"})

        self.assertContains(response, self.item.sku)
        self.assertNotContains(response, other.sku)

    def test_invalid_category_filter_does_not_crash_balances(self):
        self._receipt(self.item, "5")

        response = self.client.get("/balances/", {"category": "not-a-number", "presentation": "consolidated"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_category"], "")
        self.assertContains(response, self.item.sku)

    def test_balances_preset_consolidated_switches_presentation(self):
        response = self.client.get("/balances/", {"preset": "consolidated"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_preset"], "consolidated")
        self.assertEqual(response.context["selected_presentation"], PRESENTATION_CONSOLIDATED)
        self.assertFalse(response.context["selected_include_zero"])

    def test_balances_preset_with_zero_includes_empty_catalog_items(self):
        Item.objects.create(sku="ZERO-002", name="Пустая позиция", unit=self.unit, is_active=False)
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("5"))
        receipt.post()

        response = self.client.get("/balances/", {"preset": "with_zero"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_preset"], "with_zero")
        self.assertTrue(response.context["selected_include_zero"])
        rows = list(response.context["balances"].object_list)
        self.assertTrue(any(row["item__sku"] == "ZERO-002" for row in rows))

    def test_balance_report_shows_builtin_preset_chips(self):
        response = self.client.get("/balances/", {"preset": "nonzero"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'href="?preset=by_warehouse"', html=False)
        self.assertContains(response, 'href="?preset=consolidated"', html=False)
        self.assertContains(response, 'href="?preset=with_zero"', html=False)
        self.assertContains(response, 'href="?preset=nonzero"', html=False)
        self.assertContains(response, "По складам")
        self.assertContains(response, "Сводно")
        self.assertContains(response, "Все позиции")
        self.assertContains(response, "Только с остатком")
        self.assertContains(response, 'class="chip active"', html=False)
        self.assertContains(response, 'name="preset" value="nonzero"', html=False)

    def test_balance_preset_can_be_overridden_by_unchecked_include_zero(self):
        response = self.client.get("/balances/", {"preset": "with_zero", "include_zero": ""})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_preset"], "with_zero")
        self.assertFalse(response.context["selected_include_zero"])

    def test_balance_export_applies_preset_defaults(self):
        response = self.client.get("/export/balances.xlsx", {"preset": "consolidated"})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        metadata_sheet = workbook["Параметры"]
        metadata = {
            metadata_sheet.cell(row=row_number, column=1).value: metadata_sheet.cell(row=row_number, column=2).value
            for row_number in range(2, metadata_sheet.max_row + 1)
        }
        self.assertEqual(metadata["Представление"], "Сводно по складам")

    def test_balance_export_filters_by_category(self):
        category = ItemCategory.objects.create(name="Расходники", code="consumables")
        self.item.category = category
        self.item.save()
        other = Item.objects.create(sku="NO-CAT-XLSX", name="Без категории XLSX", unit=self.unit)
        self._receipt(self.item, "5")
        self._receipt(other, "7")

        response = self.client.get(
            "/export/balances.xlsx",
            {"category": str(category.pk), "presentation": "consolidated"},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook["Остатки"].iter_rows(min_row=2, values_only=True))
        exported_skus = {row[0] for row in rows}
        self.assertIn(self.item.sku, exported_skus)
        self.assertNotIn(other.sku, exported_skus)

    def test_invalid_category_filter_does_not_crash_balance_export(self):
        self._receipt(self.item, "5")

        response = self.client.get(
            "/export/balances.xlsx",
            {"category": "not-a-number", "presentation": "consolidated"},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook["Остатки"].iter_rows(min_row=2, values_only=True))
        self.assertIn(self.item.sku, {row[0] for row in rows})

    def test_authenticated_user_can_save_document_view(self):
        user = User.objects.create_user(username="saved-view-user", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.OPERATOR)
        self.client.force_login(user)

        response = self.client.post(
            "/saved-views/documents/create/",
            {"name": "Мои черновики", "status": DocumentStatus.DRAFT, "warehouse": str(self.warehouse.pk)},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        saved = UserSavedView.objects.get(user=user, scope="documents")
        self.assertEqual(saved.name, "Мои черновики")
        self.assertEqual(saved.query_params["status"], DocumentStatus.DRAFT)
        self.assertContains(response, "Мои черновики")

    def test_saved_view_name_is_truncated_to_field_limit(self):
        user = User.objects.create_user(username="saved-view-long-name", password="pass")
        UserProfile.objects.create(user=user, role=UserRole.OPERATOR)
        self.client.force_login(user)
        long_name = "x" * 120

        response = self.client.post(
            "/saved-views/documents/create/",
            {"name": long_name, "status": DocumentStatus.DRAFT},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        saved = UserSavedView.objects.get(user=user, scope="documents")
        self.assertEqual(len(saved.name), 80)

    def test_anonymous_user_cannot_save_view(self):
        response = self.client.post("/saved-views/documents/create/", {"name": "Test"})

        self.assertEqual(response.status_code, 302)

    def test_item_list_paginates_large_nomenclature(self):
        for index in range(30):
            Item.objects.create(sku=f"S-{index:03d}", name=f"Позиция {index:03d}", unit=self.unit)

        response = self.client.get("/items/", {"page_size": 10, "page": 2})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["page_obj"].paginator.per_page, 10)
        self.assertEqual(response.context["page_obj"].number, 2)
        self.assertEqual(len(response.context["page_obj"].object_list), 10)

    def test_document_list_supports_date_filters_and_export_link(self):
        older = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=older, item=self.item, quantity=Decimal("5"))
        older.post()

        newer = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 10),
        )
        StockDocumentLine.objects.create(document=newer, item=self.item, quantity=Decimal("-2"))
        newer.post()

        response = self.client.get(
            "/documents/",
            {"date_from": "2026-03-05", "date_to": "2026-03-31", "warehouse": str(self.warehouse.pk)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="date_from"', html=False)
        self.assertContains(response, 'name="date_to"', html=False)
        documents = list(response.context["documents"].object_list)
        self.assertEqual(len(documents), 1)
        self.assertEqual(documents[0].pk, newer.pk)
        self.assertEqual(
            response.context["movements_query_string"],
            f"warehouse={self.warehouse.pk}&date_from=2026-03-05&date_to=2026-03-31",
        )

    def test_documents_preset_drafts_filters_draft_documents(self):
        draft = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=draft, item=self.item, quantity=Decimal("5"))
        posted = self._receipt(self.item, "3")

        response = self.client.get("/documents/", {"preset": "drafts"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_preset"], "drafts")
        self.assertEqual(response.context["selected_status"], DocumentStatus.DRAFT)
        documents = list(response.context["documents"].object_list)
        self.assertIn(draft, documents)
        self.assertNotIn(posted, documents)

    def test_documents_preset_transfers_filters_transfer_documents(self):
        destination = Warehouse.objects.create(code="reserve", name="Резервный склад")
        transfer = StockDocument.objects.create(
            document_type=StockDocumentType.TRANSFER,
            warehouse=self.warehouse,
            destination_warehouse=destination,
            operation_date=timezone.localdate(),
        )
        StockDocumentLine.objects.create(document=transfer, item=self.item, quantity=Decimal("2"))
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 5),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("5"))

        response = self.client.get("/documents/", {"preset": "transfers"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_preset"], "transfers")
        self.assertEqual(response.context["selected_type"], StockDocumentType.TRANSFER)
        documents = list(response.context["documents"].object_list)
        self.assertIn(transfer, documents)
        self.assertNotIn(receipt, documents)

    def test_document_list_shows_builtin_preset_chips(self):
        response = self.client.get("/documents/", {"preset": "drafts"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'href="?preset=drafts"', html=False)
        self.assertContains(response, 'href="?preset=posted"', html=False)
        self.assertContains(response, 'href="?preset=receipts"', html=False)
        self.assertContains(response, 'href="?preset=issues"', html=False)
        self.assertContains(response, 'href="?preset=transfers"', html=False)
        self.assertContains(response, "Черновики")
        self.assertContains(response, "Проведенные")
        self.assertContains(response, "Приходы")
        self.assertContains(response, "Расходы")
        self.assertContains(response, "Перемещения")
        self.assertContains(response, 'class="chip active"', html=False)
        self.assertContains(response, 'name="preset" value="drafts"', html=False)

    def test_document_preset_allows_explicit_empty_status_override(self):
        draft = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=draft, item=self.item, quantity=Decimal("5"))
        posted = self._receipt(self.item, "3")

        response = self.client.get("/documents/", {"preset": "drafts", "status": ""})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_preset"], "drafts")
        self.assertEqual(response.context["selected_status"], "")
        documents = list(response.context["documents"].object_list)
        self.assertIn(draft, documents)
        self.assertIn(posted, documents)

    def test_document_export_link_preserves_explicit_empty_preset_override(self):
        response = self.client.get("/documents/", {"preset": "drafts", "status": ""})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["movements_query_string"], "preset=drafts&status=")
        self.assertContains(response, "/export/movements.xlsx?preset=drafts&amp;status=", html=False)

    def test_movement_export_applies_transfer_preset(self):
        receipt = self._receipt(self.item, "5")
        destination = Warehouse.objects.create(code="reserve", name="Резервный склад")
        transfer = StockDocument.objects.create(
            document_type=StockDocumentType.TRANSFER,
            warehouse=self.warehouse,
            destination_warehouse=destination,
            operation_date=timezone.localdate(),
        )
        StockDocumentLine.objects.create(document=transfer, item=self.item, quantity=Decimal("2"))
        transfer.post()

        response = self.client.get("/export/movements.xlsx", {"preset": "transfers"})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        metadata_sheet = workbook["Параметры"]
        metadata = {
            metadata_sheet.cell(row=row_number, column=1).value: metadata_sheet.cell(row=row_number, column=2).value
            for row_number in range(2, metadata_sheet.max_row + 1)
        }
        self.assertEqual(metadata["Тип документа"], "Перемещение")
        rows = list(workbook["Движения"].iter_rows(min_row=2, values_only=True))
        self.assertTrue(rows)
        self.assertEqual({row[1] for row in rows}, {transfer.number})
        self.assertNotIn(receipt.number, {row[1] for row in rows})

    def test_movement_export_applies_draft_preset(self):
        draft = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=draft, item=self.item, quantity=Decimal("5"))
        posted = self._receipt(self.item, "3")

        response = self.client.get("/export/movements.xlsx", {"preset": "drafts"})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        metadata_sheet = workbook["Параметры"]
        metadata = {
            metadata_sheet.cell(row=row_number, column=1).value: metadata_sheet.cell(row=row_number, column=2).value
            for row_number in range(2, metadata_sheet.max_row + 1)
        }
        self.assertEqual(metadata["Статус документа"], "Черновик")
        rows = list(workbook["Движения"].iter_rows(min_row=2, values_only=True))
        self.assertTrue(rows)
        self.assertEqual({row[1] for row in rows}, {draft.number})
        self.assertNotIn(posted.number, {row[1] for row in rows})

    def test_export_analysis_response_sets_download_filename(self):
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("10"))
        receipt.post()

        response = self.client.get(
            "/export/analysis.xlsx",
            {
                "mode": "month",
                "anchor_date": "2026-03-15",
                "presentation": "consolidated",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(".xlsx", response["Content-Disposition"])
        self.assertIn("analysis_2026-03-01_2026-03-31.xlsx", response["Content-Disposition"])

    def test_export_analysis_by_warehouse_creates_extra_warehouse_sheets(self):
        second_warehouse = Warehouse.objects.create(code="reserve", name="Резервный склад")

        receipt_main = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt_main, item=self.item, quantity=Decimal("10"))
        receipt_main.post()

        receipt_reserve = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=second_warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt_reserve, item=self.item, quantity=Decimal("4"))
        receipt_reserve.post()

        export_payload = export_period_analysis_xlsx(
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            label="Месяц 03.2026",
            presentation=PRESENTATION_BY_WAREHOUSE,
            mode_label="Месяц",
        )
        workbook = load_workbook(export_payload.buffer)
        self.assertIn("Склад Основной склад", workbook.sheetnames)
        self.assertIn("Склад Резервный склад", workbook.sheetnames)

        main_sheet = workbook["Склад Основной склад"]
        reserve_sheet = workbook["Склад Резервный склад"]
        self.assertEqual(main_sheet.cell(row=1, column=1).value, "Артикул")
        self.assertEqual(reserve_sheet.cell(row=1, column=1).value, "Артикул")
        self.assertEqual(main_sheet.cell(row=2, column=1).value, "A-100")
        self.assertEqual(reserve_sheet.cell(row=2, column=1).value, "A-100")

    def test_daily_and_monthly_report_views_open(self):
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("10"))
        receipt.post()

        daily_response = self.client.get("/days/", {"anchor_date": "2026-03-19", "presentation": "consolidated"})
        self.assertEqual(daily_response.status_code, 200)
        self.assertContains(daily_response, "Подневный отчет")
        self.assertGreaterEqual(daily_response.context["report"]["summary"]["rows_count"], 31)

        monthly_response = self.client.get(
            "/months/",
            {"date_from": "2026-01-01", "date_to": "2026-03-31", "presentation": "consolidated"},
        )
        self.assertEqual(monthly_response.status_code, 200)
        self.assertContains(monthly_response, "Помесячный отчет")
        self.assertGreaterEqual(monthly_response.context["report"]["summary"]["rows_count"], 3)

    def test_daily_and_monthly_reports_expose_period_navigation_queries(self):
        daily_response = self.client.get("/days/", {"anchor_date": "2026-03-19", "presentation": "consolidated"})
        self.assertEqual(daily_response.status_code, 200)
        self.assertEqual(
            daily_response.context["prev_period_query"],
            "anchor_date=2026-02-01&presentation=consolidated&page_size=50",
        )
        self.assertEqual(
            daily_response.context["next_period_query"],
            "anchor_date=2026-04-01&presentation=consolidated&page_size=50",
        )
        self.assertContains(daily_response, "?anchor_date=2026-02-01&amp;presentation=consolidated&amp;page_size=50")
        self.assertContains(daily_response, "?anchor_date=2026-04-01&amp;presentation=consolidated&amp;page_size=50")

        monthly_response = self.client.get(
            "/months/",
            {"date_from": "2026-01-17", "date_to": "2026-03-12", "presentation": "consolidated"},
        )
        self.assertEqual(monthly_response.status_code, 200)
        self.assertEqual(
            monthly_response.context["prev_period_query"],
            "date_from=2025-12-01&date_to=2026-02-28&presentation=consolidated&page_size=50",
        )
        self.assertEqual(
            monthly_response.context["next_period_query"],
            "date_from=2026-02-01&date_to=2026-04-30&presentation=consolidated&page_size=50",
        )
        self.assertContains(
            monthly_response,
            "?date_from=2025-12-01&amp;date_to=2026-02-28&amp;presentation=consolidated&amp;page_size=50",
        )
        self.assertContains(
            monthly_response,
            "?date_from=2026-02-01&amp;date_to=2026-04-30&amp;presentation=consolidated&amp;page_size=50",
        )

    def test_dashboard_uses_regular_inventory_cta_after_posted_activity(self):
        opening_response = self.client.get("/")
        self.assertEqual(opening_response.status_code, 200)
        self.assertContains(opening_response, "Стартовая инвентаризация")
        self.assertContains(opening_response, 'href="/documents/?status=draft"', html=False)
        self.assertContains(opening_response, 'href="/inventories/?status=draft"', html=False)

        self._receipt(self.item, "10")

        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Новая инвентаризация")
        self.assertNotContains(response, "Стартовая инвентаризация")

    def test_document_and_inventory_detail_add_confirm_before_posting(self):
        document = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("3"))

        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 3, 2),
            scope=InventoryScope.PARTIAL,
        )
        InventoryLine.objects.create(
            inventory=inventory,
            item=self.item,
            expected_quantity=Decimal("0"),
            actual_quantity=Decimal("2"),
        )

        document_response = self.client.get(f"/documents/{document.pk}/")
        inventory_response = self.client.get(f"/inventories/{inventory.pk}/")

        self.assertContains(document_response, "Провести документ? После проведения документ нельзя изменить.")
        self.assertContains(inventory_response, "Провести инвентаризацию? После проведения документ нельзя изменить.")
        self.assertContains(
            document_response,
            'onsubmit="return confirm(\'Провести документ? После проведения документ нельзя изменить.\');"',
            html=False,
        )
        self.assertContains(
            inventory_response,
            'onsubmit="return confirm(\'Провести инвентаризацию? После проведения документ нельзя изменить.\');"',
            html=False,
        )

    def test_document_detail_shows_activity_timeline_after_posting(self):
        document = self._receipt(self.item, "10")

        response = self.client.get(f"/documents/{document.pk}/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "История")
        self.assertContains(response, f"Документ {document.number} проведен")
        self.assertContains(response, "Приход")

    def test_inventory_detail_shows_activity_timeline_after_posting(self):
        self._receipt(self.item, "10")
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 3, 13),
            scope=InventoryScope.PARTIAL,
        )
        InventoryLine.objects.create(inventory=inventory, item=self.item, actual_quantity=Decimal("8"))

        inventory.post()
        adjustment = inventory.generated_documents.get()

        response = self.client.get(f"/inventories/{inventory.pk}/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "История")
        self.assertContains(response, f"Инвентаризация {inventory.number} проведена")
        self.assertContains(response, f"Создана автокорректировка {adjustment.number}")

    def test_document_list_has_entry_points_for_all_document_types_and_badges(self):
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("4"))

        issue = StockDocument.objects.create(
            document_type=StockDocumentType.ISSUE,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 2),
        )
        StockDocumentLine.objects.create(document=issue, item=self.item, quantity=Decimal("-1"))

        adjustment = StockDocument.objects.create(
            document_type=StockDocumentType.ADJUSTMENT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 3),
        )
        StockDocumentLine.objects.create(document=adjustment, item=self.item, quantity=Decimal("2"))

        response = self.client.get("/documents/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '?type=receipt', html=False)
        self.assertContains(response, '?type=issue', html=False)
        self.assertContains(response, '?type=adjustment', html=False)
        self.assertContains(response, 'doc-type-badge doc-type-receipt', html=False)
        self.assertContains(response, 'doc-type-badge doc-type-issue', html=False)
        self.assertContains(response, 'doc-type-badge doc-type-adjustment', html=False)

    def test_unit_list_shows_precision_column(self):
        response = self.client.get("/units/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Точность")
        self.assertContains(response, "<td>3</td>", html=True)

    def test_reference_create_and_update_record_operational_activity(self):
        admin = User.objects.create_user(username="admin-reference-audit", password="pass")
        UserProfile.objects.create(user=admin, role=UserRole.ADMIN)
        self.client.force_login(admin)

        create_response = self.client.post(
            "/units/",
            {"code": "box", "name": "Коробка", "display_precision": "0"},
        )
        self.assertRedirects(create_response, "/units/")
        unit = Unit.objects.get(code="box")

        update_response = self.client.post(
            f"/units/{unit.pk}/edit/",
            {"code": "box", "name": "Коробка обновленная", "display_precision": "0"},
        )
        self.assertRedirects(update_response, "/units/")

        events = ActivityEvent.objects.filter(event_type=ActivityEventType.REFERENCE_RECORD_CHANGED).order_by("id")
        self.assertEqual(events.count(), 2)
        self.assertEqual([event.metadata["action"] for event in events], ["created", "updated"])
        self.assertEqual([event.metadata["model"] for event in events], ["Unit", "Unit"])
        self.assertEqual(events[0].actor, admin)

    def test_document_form_shows_quantity_hint_near_table_header(self):
        response = self.client.get("/documents/new/", {"type": "adjustment"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Для корректировки можно указать знак.")
        self.assertContains(response, 'class="table-hint"', html=False)
        self.assertContains(response, f'value="{timezone.localdate().isoformat()}"', html=False)

    def test_document_form_marks_destination_warehouse_field_for_transfer_only_visibility(self):
        response = self.client.get("/documents/new/", {"type": "receipt"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-document-type-field="1"', html=False)
        self.assertContains(response, 'data-destination-warehouse-field="1"', html=False)
        self.assertContains(response, "Используется только для документа перемещения.")
        self.assertContains(response, "destinationField.hidden = !isTransfer;", html=False)
        self.assertContains(response, "typeField.value === 'transfer'", html=False)

    def test_inventory_form_uses_native_date_value_format(self):
        response = self.client.get("/inventories/new/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'value="{timezone.localdate().isoformat()}"', html=False)

    def test_document_list_supports_transfer_entry_point_and_badge(self):
        destination = Warehouse.objects.create(code="reserve", name="Резервный склад")
        transfer = StockDocument.objects.create(
            document_type=StockDocumentType.TRANSFER,
            warehouse=self.warehouse,
            destination_warehouse=destination,
            operation_date=date(2026, 3, 4),
        )
        StockDocumentLine.objects.create(document=transfer, item=self.item, quantity=Decimal("2"))

        response = self.client.get("/documents/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '?type=transfer', html=False)
        self.assertContains(response, 'doc-type-badge doc-type-transfer', html=False)

    def test_document_create_from_movements_section_preselects_and_saves_transfer(self):
        destination = Warehouse.objects.create(code="reserve", name="Резервный склад")

        response = self.client.get("/documents/new/", {"type": "transfer"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="transfer" selected', html=False)

        post_response = self.client.post(
            "/documents/new/?type=transfer",
            {
                "document_type": StockDocumentType.TRANSFER,
                "warehouse": str(self.warehouse.pk),
                "destination_warehouse": str(destination.pk),
                "operation_date": "2026-03-10",
                "comment": "Перемещение со склада",
                "lines-TOTAL_FORMS": "6",
                "lines-INITIAL_FORMS": "0",
                "lines-MIN_NUM_FORMS": "0",
                "lines-MAX_NUM_FORMS": "1000",
                "lines-0-item": str(self.item.pk),
                "lines-0-quantity": "4",
                "lines-0-comment": "",
            },
            follow=True,
        )

        self.assertEqual(post_response.status_code, 200)
        transfer = StockDocument.objects.get(document_type=StockDocumentType.TRANSFER, comment="Перемещение со склада")
        self.assertEqual(transfer.destination_warehouse, destination)
        self.assertEqual(transfer.lines.get().quantity, Decimal("4"))

    def test_draft_document_can_be_updated(self):
        destination = Warehouse.objects.create(code="reserve", name="Резервный склад")
        document = StockDocument.objects.create(
            document_type=StockDocumentType.TRANSFER,
            warehouse=self.warehouse,
            destination_warehouse=destination,
            operation_date=date(2026, 3, 5),
            comment="Черновик",
        )
        StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("2"))

        response = self.client.post(
            f"/documents/{document.pk}/edit/",
            {
                "document_type": StockDocumentType.TRANSFER,
                "warehouse": str(self.warehouse.pk),
                "destination_warehouse": str(destination.pk),
                "operation_date": "2026-03-06",
                "comment": "Обновленный черновик",
                "lines-TOTAL_FORMS": "6",
                "lines-INITIAL_FORMS": "0",
                "lines-MIN_NUM_FORMS": "0",
                "lines-MAX_NUM_FORMS": "1000",
                "lines-0-item": str(self.item.pk),
                "lines-0-quantity": "5",
                "lines-0-comment": "обновлено",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        document.refresh_from_db()
        self.assertEqual(document.comment, "Обновленный черновик")
        self.assertEqual(document.operation_date, date(2026, 3, 6))
        self.assertEqual(document.lines.count(), 1)
        self.assertEqual(document.lines.get().quantity, Decimal("5"))
        self.assertContains(response, "Черновик документа обновлен.")

    def test_posted_document_cannot_be_edited(self):
        document = self._receipt(self.item, "10")

        response = self.client.get(f"/documents/{document.pk}/edit/", follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Редактировать можно только черновик.")

    def test_draft_inventory_can_be_updated(self):
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 3, 7),
            scope=InventoryScope.PARTIAL,
            comment="Черновик инвентаризации",
        )
        InventoryLine.objects.create(
            inventory=inventory,
            item=self.item,
            actual_quantity=Decimal("2"),
            expected_quantity=Decimal("0"),
        )

        response = self.client.post(
            f"/inventories/{inventory.pk}/edit/",
            {
                "warehouse": str(self.warehouse.pk),
                "inventory_date": "2026-03-08",
                "scope": InventoryScope.FULL,
                "comment": "Обновленная инвентаризация",
                "lines-TOTAL_FORMS": "8",
                "lines-INITIAL_FORMS": "0",
                "lines-MIN_NUM_FORMS": "0",
                "lines-MAX_NUM_FORMS": "1000",
                "lines-0-item": str(self.item.pk),
                "lines-0-actual_quantity": "5",
                "lines-0-comment": "уточнено",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        inventory.refresh_from_db()
        self.assertEqual(inventory.scope, InventoryScope.FULL)
        self.assertEqual(inventory.comment, "Обновленная инвентаризация")
        self.assertEqual(inventory.lines.count(), 1)
        self.assertEqual(inventory.lines.get().actual_quantity, Decimal("5"))
        self.assertContains(response, "Черновик инвентаризации обновлен.")

    def test_transfer_document_updates_source_and_destination_balances(self):
        destination = Warehouse.objects.create(code="reserve", name="Резервный склад")
        receipt = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 1),
        )
        StockDocumentLine.objects.create(document=receipt, item=self.item, quantity=Decimal("10"))
        receipt.post()
        transfer = StockDocument.objects.create(
            document_type=StockDocumentType.TRANSFER,
            warehouse=self.warehouse,
            destination_warehouse=destination,
            operation_date=date(2026, 3, 9),
        )
        StockDocumentLine.objects.create(document=transfer, item=self.item, quantity=Decimal("4"))

        transfer.post()

        source_balance = get_balance_map(self.warehouse)[self.item.id]
        destination_balance = get_balance_map(destination)[self.item.id]
        consolidated_balance = get_balance_map(presentation=PRESENTATION_CONSOLIDATED)[self.item.id]

        self.assertEqual(source_balance, Decimal("6"))
        self.assertEqual(destination_balance, Decimal("4"))
        self.assertEqual(consolidated_balance, Decimal("10"))

        movements_export = export_movements_xlsx()
        movements_sheet = load_workbook(movements_export.buffer)["Движения"]
        rows = list(movements_sheet.iter_rows(min_row=2, values_only=True))
        transfer_rows = [row for row in rows if row[1] == transfer.number]
        self.assertEqual(len(transfer_rows), 2)
        self.assertEqual({row[3] for row in transfer_rows}, {self.warehouse.name, destination.name})
        self.assertEqual({Decimal(str(row[6])) for row in transfer_rows}, {Decimal("-4"), Decimal("4")})

    def test_stock_document_line_rejects_duplicate_item_per_document(self):
        document = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 3, 10),
        )
        StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("2"))

        with self.assertRaises(IntegrityError):
            StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("3"))

    def test_sidebar_groups_navigation_sections(self):
        response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<link rel="icon" href="data:,">', html=False)
        self.assertContains(response, "Обзор")
        self.assertContains(response, "Справочники")
        self.assertContains(response, "Документы")
        self.assertContains(response, "Отчеты")
        self.assertContains(response, 'class="nav-group-title"', html=False)

    def test_dashboard_exposes_application_version(self):
        response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, APP_VERSION_LABEL)

    def test_dashboard_exposes_local_single_user_deployment_limits(self):
        response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Local Single User")
        self.assertContains(response, "SQLite")
        self.assertContains(response, "один локальный компьютер")
        self.assertContains(response, "один активный оператор")
        self.assertContains(response, "не режим одновременной многопользовательской работы")

    def test_brand_links_to_dashboard_and_empty_states_suggest_next_action(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'href="/" class="brand"', html=False)

        self.second_item.delete()
        self.item.delete()

        items_response = self.client.get("/items/")
        self.assertContains(items_response, "Номенклатура пока не заполнена.")
        self.assertContains(items_response, "Добавьте первую позицию через форму слева")

        self.unit.delete()
        units_response = self.client.get("/units/")
        self.assertContains(units_response, "Справочник единиц пока пуст.")
        self.assertContains(units_response, "Добавьте первую единицу через форму слева")

        documents_response = self.client.get("/documents/")
        self.assertContains(documents_response, "Документы движения пока не созданы.")
        self.assertContains(documents_response, "?type=receipt", html=False)
        self.assertContains(documents_response, "?type=issue", html=False)
        self.assertContains(documents_response, "?type=adjustment", html=False)

        inventories_response = self.client.get("/inventories/")
        self.assertContains(inventories_response, "Инвентаризации пока не создавались.")
        self.assertContains(inventories_response, 'href="/inventories/new/"', html=False)

    def test_success_messages_marked_for_auto_dismiss_but_errors_are_not(self):
        session = self.client.session
        session.save()

        from django.contrib.messages.storage.fallback import FallbackStorage
        from django.contrib.messages import constants
        from django.http import HttpResponse
        from django.template.response import TemplateResponse

        request = self.client.get("/").wsgi_request
        setattr(request, "session", session)
        storage = FallbackStorage(request)
        setattr(request, "_messages", storage)
        storage.add(constants.SUCCESS, "Успешно сохранено")
        storage.add(constants.ERROR, "Ошибка сохранения")

        response = TemplateResponse(request, "base.html", {})
        response.render()
        content = response.content.decode("utf-8")

        self.assertIn('data-autodismiss="1"', content)
        self.assertIn("Ошибка сохранения", content)


class UserAttributionModelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="operator", password="pass")
        self.unit = Unit.objects.create(code="pcs", name="Штука")
        self.warehouse = Warehouse.objects.create(code="main", name="Основной склад")
        self.item = Item.objects.create(sku="SKU-1", name="Позиция", unit=self.unit)

    def test_stock_document_stores_created_updated_and_posted_users(self):
        document = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 6, 10),
            created_by=self.user,
            updated_by=self.user,
        )
        StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("2"))

        document.post(posted_by=self.user)
        document.refresh_from_db()

        self.assertEqual(document.created_by, self.user)
        self.assertEqual(document.updated_by, self.user)
        self.assertEqual(document.posted_by, self.user)

    def test_inventory_document_stores_created_updated_and_posted_users(self):
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 6, 10),
            scope=InventoryScope.FULL,
            created_by=self.user,
            updated_by=self.user,
        )
        InventoryLine.objects.create(
            inventory=inventory,
            item=self.item,
            expected_quantity=Decimal("0"),
            actual_quantity=Decimal("1"),
        )

        inventory.post(posted_by=self.user)
        inventory.refresh_from_db()

        self.assertEqual(inventory.created_by, self.user)
        self.assertEqual(inventory.updated_by, self.user)
        self.assertEqual(inventory.posted_by, self.user)

    def test_stock_document_posted_event_stores_actor(self):
        document = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 6, 10),
        )
        StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("2"))

        document.post(posted_by=self.user)

        event = ActivityEvent.objects.get(
            stock_document=document,
            event_type=ActivityEventType.STOCK_DOCUMENT_POSTED,
        )
        self.assertEqual(event.actor, self.user)
        self.assertEqual(event.actor_label, "operator")

    def test_inventory_posted_events_store_actor(self):
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 6, 10),
            scope=InventoryScope.FULL,
        )
        InventoryLine.objects.create(
            inventory=inventory,
            item=self.item,
            expected_quantity=Decimal("0"),
            actual_quantity=Decimal("1"),
        )

        inventory.post(posted_by=self.user)

        events = ActivityEvent.objects.filter(inventory_document=inventory)
        self.assertTrue(events.filter(actor=self.user, actor_label="operator").exists())


class UserAttributionViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="operator", password="pass")
        UserProfile.objects.create(user=self.user, role=UserRole.OPERATOR)
        self.unit = Unit.objects.create(code="pcs", name="Штука")
        self.warehouse = Warehouse.objects.create(code="main", name="Основной склад")
        self.item = Item.objects.create(sku="SKU-1", name="Позиция", unit=self.unit)
        self.client.force_login(self.user)

    def test_document_create_sets_created_and_updated_by(self):
        response = self.client.post(
            "/documents/new/?type=receipt",
            {
                "document_type": StockDocumentType.RECEIPT,
                "warehouse": self.warehouse.pk,
                "destination_warehouse": "",
                "operation_date": "2026-06-10",
                "comment": "",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "0",
                "lines-MIN_NUM_FORMS": "0",
                "lines-MAX_NUM_FORMS": "1000",
                "lines-0-item": self.item.pk,
                "lines-0-quantity": "2",
                "lines-0-comment": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        document = StockDocument.objects.get()
        self.assertEqual(document.created_by, self.user)
        self.assertEqual(document.updated_by, self.user)

    def test_document_post_sets_posted_by(self):
        document = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 6, 10),
        )
        StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("2"))

        response = self.client.post(f"/documents/{document.pk}/post/")

        self.assertEqual(response.status_code, 302)
        document.refresh_from_db()
        self.assertEqual(document.posted_by, self.user)

    def test_inventory_create_sets_created_and_updated_by(self):
        response = self.client.post(
            "/inventories/new/",
            {
                "warehouse": self.warehouse.pk,
                "inventory_date": "2026-06-10",
                "scope": InventoryScope.PARTIAL,
                "comment": "",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "0",
                "lines-MIN_NUM_FORMS": "0",
                "lines-MAX_NUM_FORMS": "1000",
                "lines-0-item": self.item.pk,
                "lines-0-actual_quantity": "2",
                "lines-0-comment": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        inventory = InventoryDocument.objects.get()
        self.assertEqual(inventory.created_by, self.user)
        self.assertEqual(inventory.updated_by, self.user)

    def test_inventory_post_sets_posted_by(self):
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 6, 10),
            scope=InventoryScope.FULL,
        )
        InventoryLine.objects.create(
            inventory=inventory,
            item=self.item,
            expected_quantity=Decimal("0"),
            actual_quantity=Decimal("1"),
        )

        response = self.client.post(f"/inventories/{inventory.pk}/post/")

        self.assertEqual(response.status_code, 302)
        inventory.refresh_from_db()
        self.assertEqual(inventory.posted_by, self.user)

    def test_document_detail_shows_posted_by_user(self):
        document = StockDocument.objects.create(
            document_type=StockDocumentType.RECEIPT,
            warehouse=self.warehouse,
            operation_date=date(2026, 6, 10),
            created_by=self.user,
            updated_by=self.user,
        )
        StockDocumentLine.objects.create(document=document, item=self.item, quantity=Decimal("2"))
        document.post(posted_by=self.user)

        response = self.client.get(f"/documents/{document.pk}/")

        self.assertContains(response, "Провел")
        self.assertContains(response, "operator")

    def test_inventory_detail_shows_posted_by_user(self):
        inventory = InventoryDocument.objects.create(
            warehouse=self.warehouse,
            inventory_date=date(2026, 6, 10),
            scope=InventoryScope.FULL,
            created_by=self.user,
            updated_by=self.user,
        )
        InventoryLine.objects.create(
            inventory=inventory,
            item=self.item,
            expected_quantity=Decimal("0"),
            actual_quantity=Decimal("1"),
        )
        inventory.post(posted_by=self.user)

        response = self.client.get(f"/inventories/{inventory.pk}/")

        self.assertContains(response, "Провел")
        self.assertContains(response, "operator")


@override_settings(DEMO_MODE=True)
class DemoModeTests(TestCase):
    def test_seed_demo_data_creates_sample_records(self):
        summary = seed_demo_data()

        self.assertEqual(summary["warehouses"], 3)
        self.assertEqual(summary["items"], 30)
        self.assertGreaterEqual(summary["documents"], 22)
        self.assertEqual(summary["inventories"], 2)
        self.assertGreaterEqual(summary["span_days"], 45)
        self.assertTrue(Unit.objects.filter(code="кг").exists())
        self.assertTrue(Unit.objects.filter(code="шт").exists())
        self.assertTrue(Unit.objects.filter(code="м").exists())
        self.assertTrue(Item.objects.filter(sku="100001", name="Позиция А").exists())
        self.assertTrue(Item.objects.filter(sku="100030", name="Позиция АБ").exists())
        self.assertTrue(StockDocument.objects.filter(comment__icontains="[demo]").exists())
        self.assertTrue(InventoryDocument.objects.filter(comment__icontains="[demo]").exists())

    def test_seed_demo_data_refuses_non_empty_database(self):
        unit = Unit.objects.create(code="kg", name="Килограмм")
        Warehouse.objects.create(code="main", name="Основной склад")
        Item.objects.create(sku="X-1", name="Позиция X", unit=unit)

        with self.assertRaises(ValidationError):
            seed_demo_data()

    def test_seed_demo_data_can_reset_existing_demo_dataset(self):
        first_summary = seed_demo_data()
        extra_unit = Unit.objects.get(code="кг")
        Item.objects.create(sku="Z-999", name="Временная позиция", unit=extra_unit)

        second_summary = seed_demo_data(force_reset=True)

        self.assertEqual(first_summary["warehouses"], second_summary["warehouses"])
        self.assertEqual(first_summary["items"], second_summary["items"])
        self.assertFalse(Item.objects.filter(sku="Z-999").exists())

    def test_demo_load_view_populates_empty_database(self):
        response = self.client.post("/demo/load/", {"next": "/"})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/")
        self.assertTrue(Item.objects.exists())

    def test_demo_load_view_records_operational_activity(self):
        response = self.client.post("/demo/load/", {"next": "/"})

        self.assertEqual(response.status_code, 302)
        event = ActivityEvent.objects.get(event_type=ActivityEventType.DEMO_DATA_RESET)
        self.assertEqual(event.metadata["reset_performed"], False)
        self.assertEqual(event.metadata["items"], 30)

    def test_demo_load_view_reloads_existing_demo_dataset(self):
        seed_demo_data()
        extra_unit = Unit.objects.get(code="кг")
        Item.objects.create(sku="Z-998", name="Лишняя позиция", unit=extra_unit)

        response = self.client.post("/demo/load/", {"next": "/"})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/")
        self.assertFalse(Item.objects.filter(sku="Z-998").exists())
        self.assertEqual(Item.objects.count(), 30)

    def test_demo_load_redirects_from_detail_to_dashboard_after_reset(self):
        seed_demo_data()
        document = StockDocument.objects.order_by("id").first()

        response = self.client.post("/demo/load/", {"next": f"/documents/{document.pk}/"})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/")

    def test_demo_load_ignores_external_next(self):
        response = self.client.post("/demo/load/", {"next": "https://example.com/elsewhere"})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/")

    @override_settings(DEMO_MODE=False)
    def test_demo_load_view_respects_setting(self):
        admin = User.objects.create_superuser(username="demo-admin", password="pass")
        self.client.force_login(admin)

        response = self.client.post("/demo/load/", {"next": "/"})

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Item.objects.exists())
